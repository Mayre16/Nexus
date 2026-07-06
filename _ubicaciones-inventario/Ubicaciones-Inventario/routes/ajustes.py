"""
Rutas para ajustes de inventario (inventario inicial)
Permite crear ubicaciones físicas y registrar productos con ubicación y cantidad
Aplicando Regla de Oro #4: Solo modifica StockUbicacion si la ubicación física existe
"""
from flask import Blueprint, request, jsonify, session, send_file, make_response
from routes.auth import require_auth, require_admin
from database import db
from database.models import StockUbicacion, Movimiento, Usuario, UbicacionFisica, ProductoADM, StockProductoADM, SyncLocationStatus, Discrepancia
from utils.validaciones import validar_sku, validar_ubicacion, validar_cantidad
from utils.db_helpers import db_query_with_retry
from utils.discrepancias import actualizar_discrepancias_por_skus
from utils.helpers import get_adm_client
from datetime import datetime
from sqlalchemy import func, or_
import logging
import io
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ajustes_bp = Blueprint('ajustes', __name__)
logger = logging.getLogger(__name__)


@ajustes_bp.route('/api/ajustes/ubicacion', methods=['GET'])
@require_auth
def listar_ubicaciones():
    """Lista todas las ubicaciones físicas activas del sistema (desde tabla UbicacionFisica)"""
    def _query():
        ubicaciones_fisicas = UbicacionFisica.query.filter_by(activa=True).order_by(UbicacionFisica.codigo).all()
        ubicaciones_list = [u.codigo for u in ubicaciones_fisicas]
        ubicaciones_stock = db.session.query(StockUbicacion.ubicacion).distinct().all()
        ubicaciones_stock_list = [u[0] for u in ubicaciones_stock if u[0] and u[0] not in ubicaciones_list]
        ubicaciones_list.extend(sorted(ubicaciones_stock_list))
        return ubicaciones_list

    ubicaciones_list = db_query_with_retry(_query, tag="listar_ubicaciones_ajustes")
    if ubicaciones_list is None:
        return jsonify({
            "success": False,
            "error": "Error al listar ubicaciones",
            "message": "No se pudo conectar a la base de datos. Intente nuevamente."
        }), 500
    return jsonify({
        "success": True,
        "ubicaciones": ubicaciones_list
    })


@ajustes_bp.route('/api/ajustes/ubicaciones-adm', methods=['GET'])
@require_auth
def listar_ubicaciones_adm():
    """Lista todas las ubicaciones ADM (macros) disponibles en el sistema"""
    def _query():
        ubicaciones_sync = SyncLocationStatus.query.order_by(SyncLocationStatus.location_name).all()
        ubicaciones_stock = db.session.query(
            StockProductoADM.location_id,
            StockProductoADM.location_name
        ).distinct().all()
        ubicaciones_map = {}
        for u in ubicaciones_sync:
            if u.location_id and u.location_name:
                ubicaciones_map[u.location_id] = {
                    "location_id": u.location_id,
                    "location_name": u.location_name,
                    "es_adesa": "ADESA" in u.location_name.upper()
                }
        for location_id, location_name in ubicaciones_stock:
            if location_id and location_name and location_id not in ubicaciones_map:
                ubicaciones_map[location_id] = {
                    "location_id": location_id,
                    "location_name": location_name,
                    "es_adesa": "ADESA" in location_name.upper()
                }
        ubicaciones_list = list(ubicaciones_map.values())
        ubicaciones_list.sort(key=lambda x: (not x["es_adesa"], x["location_name"]))
        return ubicaciones_list

    ubicaciones_list = db_query_with_retry(_query, tag="listar_ubicaciones_adm")
    if ubicaciones_list is None:
        return jsonify({
            "success": False,
            "error": "Error al listar ubicaciones ADM",
            "message": "No se pudo conectar a la base de datos. Intente nuevamente."
        }), 500
    return jsonify({
        "success": True,
        "ubicaciones": ubicaciones_list
    })


@ajustes_bp.route('/api/ajustes/registrar', methods=['POST'])
@require_auth
def registrar_ajuste():
    """
    Registra un ajuste de inventario (inventario inicial)
    Soporta estructura nueva (múltiples productos con asignaciones múltiples) y antigua (compatibilidad)
    Aplica Regla de Oro #4: Solo modifica StockUbicacion si la ubicación física existe y está activa
    """
    try:
        data = request.json or {}
        notas_personalizadas = data.get('notas', '').strip()  # Notas personalizadas del usuario
        
        # NUEVA ESTRUCTURA: productos con asignaciones múltiples
        productos = data.get('productos', [])
        # COMPATIBILIDAD: mantener soporte para estructura antigua
        sku_antiguo = data.get('sku', '').strip().upper()
        ubicacion_antigua = data.get('ubicacion', '').strip().upper()
        cantidad_antigua = data.get('cantidad')
        product_id_antiguo = data.get('product_id', '')
        
        # Si viene estructura nueva, usar esa. Si no, convertir estructura antigua
        if productos and len(productos) > 0:
            pass
        elif sku_antiguo and ubicacion_antigua and cantidad_antigua:
            productos = [{
                'sku': sku_antiguo,
                'item_id': product_id_antiguo,
                'cantidad_total': float(cantidad_antigua),
                'asignaciones': [{
                    'ubicacion': ubicacion_antigua,
                    'cantidad': float(cantidad_antigua)
                }]
            }]
        else:
            return jsonify({
                "success": False,
                "error": "Debe proporcionar productos con asignaciones o usar estructura antigua (sku, ubicacion, cantidad)"
            }), 400
        
        # VALIDACIÓN DE SUMATORIA POR SKU Y VALIDACIONES GENERALES
        for producto in productos:
            sku = producto.get('sku', '').strip().upper()
            cantidad_total = float(producto.get('cantidad_total', 0))
            asignaciones = producto.get('asignaciones', [])
            item_id = producto.get('item_id', '')
            
            # Validar SKU
            es_valido, mensaje = validar_sku(sku)
            if not es_valido:
                return jsonify({
                    "success": False,
                    "error": f"SKU inválido: {mensaje}"
                }), 400
            
            # Validar que haya asignaciones
            if not asignaciones or len(asignaciones) == 0:
                return jsonify({
                    "success": False,
                    "error": f"El producto {sku} debe tener al menos una asignación"
                }), 400
            
            # Calcular suma de asignaciones
            suma_asignaciones = sum(float(a.get('cantidad', 0)) for a in asignaciones)
            
            # Validar que la suma no exceda la cantidad total (si se proporciona)
            if cantidad_total > 0 and suma_asignaciones > cantidad_total:
                return jsonify({
                    "success": False,
                    "error": f"El producto {sku} tiene asignaciones que exceden la cantidad total. Total: {cantidad_total}, Suma asignada: {suma_asignaciones}"
                }), 400
            
            # Validar que todas las asignaciones tengan ubicación y cantidad válida
            for asignacion in asignaciones:
                tipo_asignacion = asignacion.get('tipo', 'fisica')  # 'fisica' o 'adm'
                cantidad = asignacion.get('cantidad', 0)
                
                # Para ajustes, permitir cantidad 0 (para eliminar stock de una ubicación)
                # Validar que sea un número válido (puede ser 0 o mayor)
                try:
                    cantidad_float = float(cantidad)
                except (ValueError, TypeError):
                    ubicacion_ref = asignacion.get('ubicacion') or asignacion.get('ubicacion_adm', '')
                    return jsonify({
                        "success": False,
                        "error": f"Cantidad inválida para {sku} en {ubicacion_ref}: Cantidad debe ser un número"
                    }), 400
                
                if cantidad_float < 0:
                    ubicacion_ref = asignacion.get('ubicacion') or asignacion.get('ubicacion_adm', '')
                    return jsonify({
                        "success": False,
                        "error": f"Cantidad inválida para {sku} en {ubicacion_ref}: Cantidad no puede ser negativa"
                    }), 400
                
                if cantidad_float > 999999.99:
                    ubicacion_ref = asignacion.get('ubicacion') or asignacion.get('ubicacion_adm', '')
                    return jsonify({
                        "success": False,
                        "error": f"Cantidad inválida para {sku} en {ubicacion_ref}: Cantidad excede el límite máximo"
                    }), 400
                
                # Si es ubicación física, validar que existe
                if tipo_asignacion == 'fisica':
                    ubicacion = asignacion.get('ubicacion', '').strip().upper()
                    
                    es_valido, mensaje = validar_ubicacion(ubicacion)
                    if not es_valido:
                        return jsonify({
                            "success": False,
                            "error": f"Ubicación física inválida para {sku}: {mensaje}"
                        }), 400
                    
                    # ✅ REGLA DE ORO #4: Validar que la ubicación física existe y está activa
                    ubicacion_fisica = UbicacionFisica.query.filter_by(
                        codigo=ubicacion,
                        activa=True
                    ).first()
                    
                    if not ubicacion_fisica:
                        return jsonify({
                            "success": False,
                            "error": f"La ubicación física '{ubicacion}' no existe o está inactiva. Verifique que la ubicación esté creada en el sistema."
                        }), 400
        
        # Procesar asignaciones
        movimientos_creados = []
        skus_ajustados_fisica = set()
        timestamp_ajuste = datetime.utcnow()  # Mismo timestamp para agrupar ajustes
        
        for producto in productos:
            sku = producto.get('sku', '').strip().upper()
            item_id = producto.get('item_id', '')
            asignaciones = producto.get('asignaciones', [])
            
            # Buscar producto usando ItemID (prioridad) con fallback a SKU
            from utils.helpers import resolver_producto_adm
            producto_db = resolver_producto_adm(item_id=item_id, sku=sku)
            
            # Si no se proporciona item_id, usar el del producto_db o buscar en ADM Cloud
            if not item_id:
                if producto_db:
                    item_id = producto_db.item_id
                else:
                    try:
                        adm_client = get_adm_client()
                        producto_adm = adm_client.buscar_item_por_sku(sku)
                        if producto_adm:
                            item_id = producto_adm.get("ID", "")
                        else:
                            item_id = sku
                    except Exception as e:
                        logger.warning(f"No se pudo obtener product_id de ADM Cloud para SKU {sku}: {e}")
                        item_id = sku
            
            for asignacion in asignaciones:
                tipo_asignacion = asignacion.get('tipo', 'fisica')
                cantidad_nueva = float(asignacion.get('cantidad', 0))
                
                if tipo_asignacion == 'fisica':
                    # Ajuste de ubicación física (dentro de ADESA)
                    ubicacion = asignacion.get('ubicacion', '').strip().upper()
                    
                    # ✅ REGLA DE ORO #4: Verificar que la ubicación física existe (ya validado arriba)
                    ubicacion_fisica = UbicacionFisica.query.filter_by(
                        codigo=ubicacion,
                        activa=True
                    ).first()
                    
                    if ubicacion_fisica:
                        skus_ajustados_fisica.add(sku)
                        # Buscar stock existente en esta ubicación usando product_id (corrige error de duplicado)
                        stock_ubic = StockUbicacion.query.filter_by(
                            product_id=item_id,
                            ubicacion=ubicacion
                        ).first()
                        
                        cantidad_anterior = 0
                        if stock_ubic:
                            cantidad_anterior = float(stock_ubic.cantidad)
                            stock_ubic.cantidad = cantidad_nueva
                            stock_ubic.updated_at = datetime.utcnow()
                        else:
                            # Usar sku de producto_db si existe (consistencia con ProductoADM); si no, sku del request
                            sku_guardar = producto_db.sku if producto_db else sku
                            stock_ubic = StockUbicacion(
                                product_id=item_id,
                                sku=sku_guardar,
                                ubicacion=ubicacion,
                                cantidad=cantidad_nueva,
                                updated_at=datetime.utcnow()
                            )
                            db.session.add(stock_ubic)
                        
                        # Calcular diferencia
                        diferencia = cantidad_nueva - cantidad_anterior
                        
                        # Crear movimiento solo si hay diferencia
                        if diferencia != 0:
                            notas_movimiento = f"Ajuste de inventario. Ubicación física: {ubicacion}. Anterior: {cantidad_anterior}, Nuevo: {cantidad_nueva}"
                            if notas_personalizadas:
                                notas_movimiento = f"{notas_personalizadas}. {notas_movimiento}"
                            
                            movimiento = Movimiento(
                                tipo="ADJUSTMENT",
                                product_id=item_id,
                                sku=sku,
                                ubicacion_origen=ubicacion if diferencia < 0 else None,
                                ubicacion_destino=ubicacion if diferencia > 0 else None,
                                cantidad=abs(diferencia),
                                usuario_id=session.get('user_id'),
                                timestamp=timestamp_ajuste,
                                notas=notas_movimiento
                            )
                            db.session.add(movimiento)
                            movimientos_creados.append(movimiento.to_dict())
                
                elif tipo_asignacion == 'adm':
                    # Ajuste de ubicación ADM (macro, no-ADESA)
                    ubicacion_adm = asignacion.get('ubicacion_adm', '')
                    location_id = asignacion.get('location_id', '')
                    
                    # Obtener stock actual en ADM (desde cache)
                    # Mejorado: Intentar obtener stock incluso si producto_db no existe
                    stock_adm_actual = 0
                    if location_id:
                        # ✅ STAGING: Obtener stock vigente (LIVE) usando helper
                        from utils.helpers import obtener_stock_vigente
                        
                        if producto_db:
                            stock_adm = obtener_stock_vigente(producto_db.id, location_id)
                            logger.info(f"[AJUSTE ADM] Buscando stock vigente para SKU {sku}, producto_db.id={producto_db.id}, location_id={location_id}, encontrado={stock_adm is not None}")
                        elif item_id:
                            # Método 2: Si no hay producto_db, intentar buscar producto por item_id primero
                            producto_por_item_id = ProductoADM.query.filter_by(item_id=item_id).first()
                            if producto_por_item_id:
                                stock_adm = obtener_stock_vigente(producto_por_item_id.id, location_id)
                                logger.info(f"[AJUSTE ADM] Buscando stock vigente para SKU {sku}, producto_por_item_id.id={producto_por_item_id.id}, location_id={location_id}, encontrado={stock_adm is not None}")
                            else:
                                stock_adm = None
                                logger.warning(f"[AJUSTE ADM] No se encontró producto por item_id={item_id} para SKU {sku}")
                        else:
                            stock_adm = None
                            logger.warning(f"[AJUSTE ADM] No hay location_id o item_id para buscar stock. SKU={sku}, location_id={location_id}, item_id={item_id}")
                        
                        if stock_adm:
                            stock_adm_actual = float(stock_adm.stock) if stock_adm.stock else 0
                            logger.info(f"[AJUSTE ADM] Stock vigente encontrado para {ubicacion_adm}: {stock_adm_actual}")
                        else:
                            logger.warning(f"[AJUSTE ADM] No se encontró stock vigente para {ubicacion_adm} (SKU={sku}, location_id={location_id}). Usando stock_adm_actual=0")
                    else:
                        logger.warning(f"[AJUSTE ADM] No hay location_id para {ubicacion_adm} (SKU={sku})")
                    
                    # Calcular diferencia
                    diferencia = cantidad_nueva - stock_adm_actual
                    logger.info(f"[AJUSTE ADM] Calculo diferencia: cantidad_nueva={cantidad_nueva}, stock_adm_actual={stock_adm_actual}, diferencia={diferencia}")
                    
                    # Solo crear movimiento si hay diferencia
                    if diferencia != 0:
                        # Actualizar cache ADM via helper centralizado (asigna sync_run_id correcto)
                        from utils.helpers import actualizar_cache_adm
                        _ant, _nue, fila_creada = actualizar_cache_adm(
                            producto_id=producto_db.id if producto_db else None,
                            location_id=location_id,
                            valor_absoluto=cantidad_nueva,
                            location_name=ubicacion_adm
                        ) if (producto_db and location_id) else (stock_adm_actual, cantidad_nueva, False)
                        
                        # Notas con Anterior/Nuevo para reversión; marcar si se creó fila nueva
                        notas_movimiento = f"Ajuste de inventario. Ubicación ADM: {ubicacion_adm}. Anterior: {stock_adm_actual}, Nuevo: {cantidad_nueva}"
                        if fila_creada:
                            notas_movimiento += " [FILA_CREADA]"
                        if notas_personalizadas:
                            notas_movimiento = f"{notas_personalizadas}. {notas_movimiento}"
                        
                        movimiento = Movimiento(
                            tipo="ADJUSTMENT",
                            product_id=item_id,
                            sku=sku,
                            ubicacion_origen=ubicacion_adm if diferencia < 0 else None,
                            ubicacion_destino=ubicacion_adm if diferencia > 0 else None,
                            cantidad=abs(diferencia),
                            usuario_id=session.get('user_id'),
                            timestamp=timestamp_ajuste,
                            notas=notas_movimiento
                        )
                        db.session.add(movimiento)
                        movimientos_creados.append(movimiento.to_dict())
                        logger.info(f"[AJUSTE ADM] Movimiento creado: {ubicacion_adm}, SKU={sku}, diferencia={diferencia}, cantidad={abs(diferencia)}")
                    else:
                        # Log informativo: no se requiere ajuste porque ya está en el valor deseado
                        logger.info(f"[AJUSTE ADM] Ajuste a {ubicacion_adm} para SKU {sku} no requiere cambio (stock actual: {stock_adm_actual}, ajuste a: {cantidad_nueva}, diferencia=0)")
        
        db.session.commit()

        # Actualizar/resolver discrepancias pendientes tras ajuste físico
        actualizar_discrepancias_por_skus(skus_ajustados_fisica)
        
        # Mensaje más descriptivo según si se crearon movimientos o no
        if len(movimientos_creados) > 0:
            mensaje = "Ajuste registrado exitosamente"
        else:
            mensaje = "No se crearon movimientos. Verifica que haya diferencia entre el stock actual y el ajuste deseado."
        
        return jsonify({
            "success": True,
            "message": mensaje,
            "movimientos": movimientos_creados,
            "total_movimientos": len(movimientos_creados),
            "advertencia": "No se crearon movimientos" if len(movimientos_creados) == 0 else None
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al registrar ajuste: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "Error al registrar ajuste",
            "message": str(e)
        }), 500


@ajustes_bp.route('/api/ajustes/buscar-producto', methods=['POST', 'OPTIONS'])
def buscar_producto_ajuste():
    """
    Busca un producto por SKU, nombre o código de barras para usar en ajustes
    USA BASE DE DATOS LOCAL (cache) para búsquedas rápidas, igual que Consulta de Productos
    """
    # Manejar preflight OPTIONS (antes del require_auth)
    if request.method == 'OPTIONS':
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response
    
    # Requerir autenticación solo para POST
    if not session.get('user_id'):
        return jsonify({
            "success": False,
            "error": "Autenticación requerida"
        }), 401
    
    try:
        data = request.json or {}
        busqueda = data.get('busqueda', '').strip()
        tipo_busqueda = data.get('tipo', 'sku').lower()  # sku, nombre, codigo_barras
        
        if not busqueda:
            return jsonify({
                "success": False,
                "error": "Término de búsqueda requerido"
            }), 400
        
        # Verificar si hay productos en cache
        total_cache = ProductoADM.query.count()
        if total_cache == 0:
            return jsonify({
                "success": False,
                "error": "Base de datos de productos no sincronizada",
                "message": "Por favor, sincroniza los productos desde ADM Cloud primero"
            }), 400
        
        producto_db = None
        
        # Buscar en base de datos local según el tipo de búsqueda (igual que productos.py)
        if tipo_busqueda == 'sku':
            from utils.helpers import resolver_producto_adm
            producto_db = resolver_producto_adm(sku=busqueda.upper())
        
        elif tipo_busqueda == 'codigo_barras':
            # Búsqueda por código de barras
            busqueda_upper = busqueda.strip().upper()
            producto_db = ProductoADM.query.filter(
                ProductoADM.codigo_barras.ilike(busqueda_upper)
            ).first()
            
            # Si no se encuentra exacto, buscar normalizando espacios y guiones
            if not producto_db:
                busqueda_normalizada = busqueda_upper.replace('-', '').replace(' ', '').replace('_', '')
                from sqlalchemy import func
                productos = ProductoADM.query.filter(
                    ProductoADM.codigo_barras.isnot(None),
                    func.length(ProductoADM.codigo_barras) >= len(busqueda_normalizada) - 3,
                    func.length(ProductoADM.codigo_barras) <= len(busqueda_normalizada) + 3
                ).all()
                for p in productos:
                    if p.codigo_barras:
                        codigo_normalizado = p.codigo_barras.replace('-', '').replace(' ', '').replace('_', '').upper()
                        if codigo_normalizado == busqueda_normalizada:
                            producto_db = p
                            break
        
        elif tipo_busqueda == 'nombre':
            # Búsqueda parcial por nombre
            busqueda_lower = busqueda.lower()
            producto_db = ProductoADM.query.filter(
                ProductoADM.nombre.ilike(f'%{busqueda_lower}%')
            ).first()
        
        if not producto_db:
            return jsonify({
                "success": False,
                "error": f"Producto no encontrado con {tipo_busqueda}: {busqueda}",
                "sugerencia": "Verifica el término de búsqueda o sincroniza los productos desde ADM Cloud"
            }), 404
        
        # Convertir ProductoADM a formato similar al de ADM Cloud para compatibilidad
        producto_encontrado = {
            "ID": producto_db.item_id,
            "SKU": producto_db.sku,
            "ItemSKU": producto_db.sku,
            "Name": producto_db.nombre or producto_db.sku,
            "BarCode": producto_db.codigo_barras
        }
        
        # ✅ STAGING: Obtener stock vigente (LIVE) para todas las ubicaciones
        from utils.helpers import obtener_stock_vigente
        from database.models import SyncLocationStatus
        
        estados_sync = SyncLocationStatus.query.filter(
            SyncLocationStatus.current_run_id.isnot(None)
        ).all()
        
        stock_ubicaciones_adm = []
        for estado in estados_sync:
            stock_vigente = obtener_stock_vigente(producto_db.id, estado.location_id)
            if stock_vigente:
                stock_ubicaciones_adm.append(stock_vigente)
        
        # Fallback: si no hay ubicaciones con current_run_id, usar registros sin sync_run_id (migración gradual)
        if not stock_ubicaciones_adm:
            stock_ubicaciones_adm = StockProductoADM.query.filter_by(
                producto_id=producto_db.id,
                sync_run_id=None
            ).all()
        
        ubicaciones_adm = []
        stock_total_adm = 0.0
        fecha_actualizacion_adm = None
        
        for stock_adm in stock_ubicaciones_adm:
            stock_cantidad = float(stock_adm.stock) if stock_adm.stock else 0.0
            
            # Solo incluir ubicaciones con stock > 0
            if stock_cantidad > 0:
                ubicaciones_adm.append({
                    "nombre": stock_adm.location_name,
                    "id": stock_adm.location_id,
                    "stock": stock_cantidad,
                    "updated_at": stock_adm.updated_at.isoformat() if stock_adm.updated_at else None
                })
                stock_total_adm += stock_cantidad
                
                # Obtener la fecha de actualización más reciente
                if stock_adm.updated_at:
                    if fecha_actualizacion_adm is None or stock_adm.updated_at > fecha_actualizacion_adm:
                        fecha_actualizacion_adm = stock_adm.updated_at
        
        # Ordenar por stock descendente (ADESA primero si tiene stock)
        ubicaciones_adm.sort(key=lambda x: (x["nombre"].upper() != "ADESA", -x["stock"]))
        
        # Obtener stock actual en todas las ubicaciones físicas WMS
        # Usar product_id (item_id) en lugar de sku: evita fallos por inconsistencia SKU (ej: "CT-5" vs "CT5")
        stock_ubicaciones = StockUbicacion.query.filter_by(
            product_id=producto_db.item_id
        ).all()
        
        ubicaciones_stock = [s.to_dict() for s in stock_ubicaciones]
        
        return jsonify({
            "success": True,
            "producto": producto_encontrado,
            "stock_ubicaciones": ubicaciones_stock,
            "stock_adm": {
                "total": stock_total_adm,
                "ubicaciones": ubicaciones_adm,
                "fecha_actualizacion": fecha_actualizacion_adm.isoformat() if fecha_actualizacion_adm else None
            }
        })
            
    except Exception as e:
        logger.error(f"Error al buscar producto para ajuste: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "Error al buscar producto",
            "message": str(e)
        }), 500


MAX_FILAS_WMS = 10000
MAX_PRODUCTOS_COMPLETO = 7500


@ajustes_bp.route('/api/ajustes/catalogo-info', methods=['GET'])
@require_auth
def catalogo_info():
    """
    Devuelve conteos para ambos tipos de catálogo.
    Usado en el modal para mostrar "X de Y" y advertencias de truncamiento.
    """
    def _count_wms():
        q = db.session.query(StockUbicacion, ProductoADM).join(
            ProductoADM, StockUbicacion.product_id == ProductoADM.item_id
        ).filter(StockUbicacion.cantidad > 0, ProductoADM.activo == True)
        total = q.count()
        return min(total, MAX_FILAS_WMS), total, total > MAX_FILAS_WMS

    def _count_completo():
        total = ProductoADM.query.filter_by(activo=True).count()
        descargados = min(total, MAX_PRODUCTOS_COMPLETO)
        return descargados, total, total > MAX_PRODUCTOS_COMPLETO

    try:
        wms_descargados, wms_total, wms_truncado = _count_wms()
        compl_descargados, compl_total, compl_truncado = _count_completo()
        return jsonify({
            "success": True,
            "wms": {
                "descargados": wms_descargados,
                "total": wms_total,
                "truncado": wms_truncado,
                "limite": MAX_FILAS_WMS,
                "mensaje": f"{wms_descargados:,} de {wms_total:,} filas" + (" (quedan excluidas)" if wms_truncado else " (completo)")
            },
            "completo": {
                "descargados": compl_descargados,
                "total": compl_total,
                "truncado": compl_truncado,
                "limite": MAX_PRODUCTOS_COMPLETO,
                "mensaje": f"{compl_descargados:,} de {compl_total:,} productos" + (" (límite por descarga)" if compl_truncado else " (completo)")
            }
        })
    except Exception as e:
        logger.error(f"Error al obtener info catálogo: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@ajustes_bp.route('/api/ajustes/descargar-catalogo', methods=['GET'])
@require_auth
def descargar_catalogo():
    """
    Descarga catálogo según tipo:
    - tipo=wms (default): Solo productos con stock > 0 en ubicaciones físicas. Para ajustes rápidos.
    - tipo=completo: Todos los productos activos con stock ref. ADM. Para trabajo desde cero.
    """
    if not HAS_OPENPYXL:
        return jsonify({
            "success": False,
            "error": "Librería openpyxl no instalada. Instala con: pip install openpyxl"
        }), 500

    tipo = request.args.get('tipo', 'wms').lower()
    if tipo == 'completo':
        return _generar_catalogo_completo()
    return _generar_catalogo_wms()


def _generar_catalogo_wms():
    """Catálogo solo con productos que tienen stock en ubicaciones físicas (ajustes rápidos)."""
    try:
        stock_con_productos = db.session.query(
            StockUbicacion,
            ProductoADM
        ).join(
            ProductoADM,
            StockUbicacion.product_id == ProductoADM.item_id
        ).filter(
            StockUbicacion.cantidad > 0,
            ProductoADM.activo == True
        ).order_by(
            ProductoADM.sku,
            StockUbicacion.ubicacion
        ).limit(MAX_FILAS_WMS + 1).all()

        total_filas = len(stock_con_productos)
        truncado = total_filas > MAX_FILAS_WMS
        if truncado:
            stock_con_productos = stock_con_productos[:MAX_FILAS_WMS]
            logger.warning(f"Catálogo WMS truncado a {MAX_FILAS_WMS} filas")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo Ajustes WMS"

        headers = ['Product ID', 'SKU', 'Nombre', 'Ubicación Física', 'Cantidad Actual']
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for stock_ubic, producto in stock_con_productos:
            ws.append([
                producto.item_id,
                producto.sku or '',
                producto.nombre or '',
                stock_ubic.ubicacion,
                float(stock_ubic.cantidad) if stock_ubic.cantidad else 0.0
            ])

        # Hoja Resumen (siempre: usuario ve X de Y)
        total_real = total_filas if not truncado else db.session.query(StockUbicacion.id).join(
            ProductoADM, StockUbicacion.product_id == ProductoADM.item_id
        ).filter(StockUbicacion.cantidad > 0, ProductoADM.activo == True).count()
        descargados_wms = len(stock_con_productos)
        ws_res = wb.create_sheet("Resumen", 0)
        ws_res.append(["RESUMEN DE DESCARGA"])
        if truncado:
            ws_res.append([f"Descargado: {descargados_wms:,} de {total_real:,} filas (productos con stock en ubicaciones físicas)"])
            ws_res.append([f"Quedan {total_real - descargados_wms:,} filas excluidas por límite de descarga."])
        else:
            ws_res.append([f"Descargado: {descargados_wms:,} de {total_real:,} filas (completo)."])
        ws_res.column_dimensions['A'].width = 70

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="catalogo_ajustes_wms_{fecha}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Error al generar catálogo WMS: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


def _generar_catalogo_completo():
    """
    Catálogo completo: todos los productos activos con stock ref. ADM.
    Incluye hoja de instrucciones. Para trabajar desde cero asignando ubicaciones físicas.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        # Stock ADM vigente por producto (ADESA como referencia) - usar run vigente si existe
        stock_adm_por_producto = {}
        estado_adesa = SyncLocationStatus.query.filter(
            SyncLocationStatus.location_name.ilike('%ADESA%')
        ).first()
        if estado_adesa and estado_adesa.current_run_id:
            stock_adm_rows = db.session.query(
                StockProductoADM.producto_id,
                StockProductoADM.stock
            ).filter(
                StockProductoADM.location_id == estado_adesa.location_id,
                StockProductoADM.sync_run_id == estado_adesa.current_run_id
            ).all()
        else:
            # Fallback legacy: stock ADESA sin sync_run_id
            stock_adm_rows = db.session.query(
                StockProductoADM.producto_id,
                StockProductoADM.stock
            ).filter(
                StockProductoADM.location_name.ilike('%ADESA%'),
                StockProductoADM.sync_run_id.is_(None)
            ).all()
        for r in stock_adm_rows:
            stock_adm_por_producto[r.producto_id] = float(r.stock) if r.stock else 0.0

        # Productos activos
        total_productos = ProductoADM.query.filter_by(activo=True).count()
        productos = ProductoADM.query.filter_by(activo=True).order_by(ProductoADM.sku).limit(MAX_PRODUCTOS_COMPLETO + 1).all()
        truncado_completo = len(productos) > MAX_PRODUCTOS_COMPLETO
        if truncado_completo:
            productos = productos[:MAX_PRODUCTOS_COMPLETO]
            logger.warning(f"Catálogo completo truncado a {MAX_PRODUCTOS_COMPLETO} productos")

        # Stock WMS por producto (product_id = item_id, ubicacion -> cantidad)
        stock_wms = db.session.query(
            StockUbicacion.product_id,
            StockUbicacion.ubicacion,
            StockUbicacion.cantidad
        ).filter(StockUbicacion.cantidad > 0).all()
        wms_por_producto = {}
        for s in stock_wms:
            key = (s.product_id, s.ubicacion)
            wms_por_producto[key] = float(s.cantidad) if s.cantidad else 0.0

        wb = openpyxl.Workbook()

        # Hoja 1: Catálogo (siempre primera para que la carga masiva lea los datos)
        ws_cat = wb.active
        ws_cat.title = "Catálogo"
        headers = ['Product ID', 'SKU', 'Nombre', 'Stock Ref. ADM', 'Ubicación Física', 'Cantidad Actual']
        ws_cat.append(headers)

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_cat[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        filas_agregadas = 0
        for producto in productos:
            stock_ref_adm = stock_adm_por_producto.get(producto.id, 0.0)

            # Productos con stock en ubicaciones físicas: una fila por ubicación
            filas_producto = [(u, c) for (pid, u), c in wms_por_producto.items() if pid == producto.item_id]
            if filas_producto:
                for ubicacion, cantidad in filas_producto:
                    ws_cat.append([
                        producto.item_id,
                        producto.sku or '',
                        producto.nombre or '',
                        stock_ref_adm,
                        ubicacion,
                        cantidad
                    ])
                    filas_agregadas += 1
            else:
                # Producto sin ubicaciones físicas: una fila para que el usuario asigne
                ws_cat.append([
                    producto.item_id,
                    producto.sku or '',
                    producto.nombre or '',
                    stock_ref_adm,
                    "",  # Usuario completa ubicación
                    0.0   # Usuario completa cantidad
                ])
                filas_agregadas += 1

        ws_cat.column_dimensions['A'].width = 40
        ws_cat.column_dimensions['B'].width = 20
        ws_cat.column_dimensions['C'].width = 50
        ws_cat.column_dimensions['D'].width = 18
        ws_cat.column_dimensions['E'].width = 20
        ws_cat.column_dimensions['F'].width = 18

        # Hoja 2: Instrucciones (segunda para no interferir con la carga masiva)
        ws_inst = wb.create_sheet("Instrucciones", 1)
        descargados_completo = len(productos)
        inst_rows = [
            ["INSTRUCCIONES - CATÁLOGO COMPLETO PARA AJUSTES"],
            [""],
        ]
        if truncado_completo:
            inst_rows.extend([
                [f"RESUMEN: Descargado {descargados_completo:,} de {total_productos:,} productos activos (límite por descarga)."],
                [f"Quedan {total_productos - descargados_completo:,} productos excluidos."],
                [""],
            ])
        else:
            inst_rows.extend([[f"RESUMEN: Descargado {descargados_completo:,} de {total_productos:,} productos (completo)."], [""]])
        inst_rows.extend([
            ["COLUMNAS A LLENAR PARA LA CARGA MASIVA:"],
            ["• Product ID (A): Identificador único. NO modificar."],
            ["• SKU (B): Código del producto. NO modificar."],
            ["• Nombre (C): Nombre del producto. NO modificar."],
            ["• Stock Ref. ADM (D): Stock actual en ADM (referencia). NO modificar. Solo información."],
            ["• Ubicación Física (E): Código de la ubicación (ej: TIENDA, 1L1AN1). OBLIGATORIO para carga."],
            ["• Cantidad Actual (F): Cantidad en esa ubicación. Use 0 para dejar vacía."],
            [""],
            ["REGLA: UN PRODUCTO EN MÚLTIPLES UBICACIONES = UNA FILA POR CADA UBICACIÓN"],
            ["  Ejemplo: Producto X en TIENDA (10) y 1L1AN1 (5) → dos filas:"],
            ["    Fila 1: Product X | ... | TIENDA  | 10"],
            ["    Fila 2: Product X | ... | 1L1AN1  | 5"],
            [""],
            ["PRODUCTOS SIN UBICACIÓN: Complete Ubicación Física y Cantidad en la fila existente."],
            ["  La ubicación debe existir en el sistema (Ubicaciones Físicas WMS)."],
            [""],
            ["Al terminar de editar, use 'Carga Masiva' para subir el archivo."],
        ])
        for row in inst_rows:
            ws_inst.append(row)
        ws_inst.column_dimensions['A'].width = 70

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="catalogo_completo_{fecha}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Error al generar catálogo completo: {str(e)}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@ajustes_bp.route('/api/ajustes/cargar-excel', methods=['POST', 'OPTIONS'])
def cargar_ajustes_excel():
    """
    Procesa un archivo Excel con ajustes masivos de inventario
    Formato esperado:
    - Columna A: SKU
    - Columna B: Ubicación Física
    - Columna C: Cantidad
    - Columna D: Notas (opcional)
    - Columna E: Product ID (opcional)
    """
    # Manejar preflight OPTIONS
    if request.method == 'OPTIONS':
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response
    
    # Requerir autenticación
    if not session.get('user_id'):
        return jsonify({
            "success": False,
            "error": "Autenticación requerida"
        }), 401
    
    if not HAS_OPENPYXL:
        return jsonify({
            "success": False,
            "error": "Librería openpyxl no instalada. Instala con: pip install openpyxl"
        }), 500
    
    try:
        # Verificar que se envió un archivo
        if 'archivo' not in request.files:
            return jsonify({
                "success": False,
                "error": "No se envió ningún archivo"
            }), 400
        
        archivo = request.files['archivo']
        notas_generales = request.form.get('notas', '').strip()  # Notas generales del ajuste
        
        if archivo.filename == '':
            return jsonify({
                "success": False,
                "error": "No se seleccionó ningún archivo"
            }), 400
        
        # Verificar extensión
        if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({
                "success": False,
                "error": "El archivo debe ser Excel (.xlsx o .xls)"
            }), 400
        
        archivo.seek(0, 2)
        file_size = archivo.tell()
        archivo.seek(0)
        max_size = 10 * 1024 * 1024  # 10 MB
        if file_size > max_size:
            return jsonify({
                "success": False,
                "error": f"Archivo demasiado grande ({file_size // 1024}KB). Máximo permitido: 10MB"
            }), 400
        
        # Leer archivo Excel
        try:
            wb = openpyxl.load_workbook(io.BytesIO(archivo.read()))
            ws = wb.active
        except Exception as e:
            return jsonify({
                "success": False,
                "error": f"Error al leer archivo Excel: {str(e)}"
            }), 400
        
        # Leer encabezados (primera fila)
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Validar que tenga las columnas mínimas
        if len(headers) < 3:
            return jsonify({
                "success": False,
                "error": "El archivo debe tener al menos 3 columnas: SKU, Ubicación Física, Cantidad"
            }), 400
        
        # Mapear columnas (flexible: buscar por nombre o posición)
        # ✅ NUEVO FORMATO: Product ID, SKU, Nombre, Ubicación, Cantidad Actual
        # ✅ FORMATO ANTIGUO: SKU, Ubicación, Cantidad (compatibilidad)
        col_sku = None
        col_ubicacion = None
        col_cantidad = None
        col_notas = None
        col_product_id = None
        col_nombre = None
        
        # Detectar formato: si tiene "Product ID" es el nuevo formato
        es_nuevo_formato = False
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if ('product' in header_lower and 'id' in header_lower) and col_product_id is None:
                col_product_id = idx
                es_nuevo_formato = True
            elif 'sku' in header_lower and col_sku is None:
                col_sku = idx
            elif ('ubicacion' in header_lower or 'ubicación' in header_lower) and col_ubicacion is None:
                col_ubicacion = idx
            elif ('cantidad' in header_lower and 'actual' in header_lower) and col_cantidad is None:
                col_cantidad = idx
            elif 'cantidad' in header_lower and col_cantidad is None:
                col_cantidad = idx
            elif 'nota' in header_lower and col_notas is None:
                col_notas = idx
            elif 'nombre' in header_lower and col_nombre is None:
                col_nombre = idx
        
        # Si no se encuentran por nombre, usar posición según formato
        if es_nuevo_formato:
            # Nuevo formato: Product ID (A), SKU (B), Nombre (C), Ubicación (D), Cantidad Actual (E)
            if col_product_id is None:
                col_product_id = 0
            if col_sku is None:
                col_sku = 1
            if col_nombre is None:
                col_nombre = 2
            if col_ubicacion is None:
                col_ubicacion = 3
            if col_cantidad is None:
                col_cantidad = 4
        else:
            # Formato antiguo: SKU (A), Ubicación (B), Cantidad (C)
            if col_sku is None:
                col_sku = 0
            if col_ubicacion is None:
                col_ubicacion = 1
            if col_cantidad is None:
                col_cantidad = 2
            if col_notas is None and len(headers) > 3:
                col_notas = 3
            if col_product_id is None and len(headers) > 4:
                col_product_id = 4
        
        # Leer datos (desde la fila 2 en adelante)
        MAX_FILAS_PERMITIDAS = 6000
        total_filas_excel = ws.max_row - 1  # Excluir header
        
        if total_filas_excel > MAX_FILAS_PERMITIDAS:
            return jsonify({
                "success": False,
                "error": f"El archivo Excel tiene {total_filas_excel} filas. El límite máximo es {MAX_FILAS_PERMITIDAS} filas por carga masiva. Por favor, divide el archivo en partes más pequeñas."
            }), 400
        
        ajustes = []
        errores = []
        timestamp_ajuste = datetime.utcnow()  # Mismo timestamp para agrupar
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Saltar filas vacías
            if not any(cell.value for cell in row):
                continue
            
            try:
                # ✅ NUEVO FORMATO: Leer Product ID primero
                product_id_raw = row[col_product_id].value if col_product_id is not None and col_product_id < len(row) else None
                product_id = str(product_id_raw).strip() if product_id_raw else ''
                
                # Leer SKU (para validación o formato antiguo)
                sku_raw = row[col_sku].value if col_sku is not None and col_sku < len(row) else None
                if sku_raw is not None:
                    if isinstance(sku_raw, (int, float)):
                        sku = str(int(float(sku_raw))).strip().upper()
                    else:
                        sku = str(sku_raw).strip().upper()
                else:
                    sku = ''
                
                ubicacion = str(row[col_ubicacion].value).strip().upper() if col_ubicacion is not None and col_ubicacion < len(row) and row[col_ubicacion].value else ''
                cantidad_str = row[col_cantidad].value if col_cantidad is not None and col_cantidad < len(row) else None
                notas_fila = str(row[col_notas].value).strip() if col_notas is not None and col_notas < len(row) and row[col_notas].value else ''
                
                # Validar campos obligatorios
                if es_nuevo_formato:
                    # Nuevo formato: Product ID es obligatorio
                    if not product_id:
                        errores.append({
                            "fila": row_idx,
                            "error": "Product ID vacío (requerido en formato catálogo)"
                        })
                        continue
                else:
                    # Formato antiguo: SKU es obligatorio
                    if not sku:
                        errores.append({
                            "fila": row_idx,
                            "error": "SKU vacío"
                        })
                        continue
                
                if not ubicacion:
                    errores.append({
                        "fila": row_idx,
                        "error": "Ubicación física vacía"
                    })
                    continue
                
                # Convertir cantidad
                try:
                    cantidad = float(cantidad_str) if cantidad_str else 0.0
                except (ValueError, TypeError):
                    errores.append({
                        "fila": row_idx,
                        "error": f"Cantidad inválida: {cantidad_str}"
                    })
                    continue
                
                # Validar cantidad: permitir 0 para ajustes (ej: dejar ubicación en 0 al mover stock)
                if cantidad < 0:
                    errores.append({
                        "fila": row_idx,
                        "error": "Cantidad no puede ser negativa"
                    })
                    continue
                
                # Resolver producto: ItemID tiene prioridad sobre SKU
                from utils.helpers import resolver_producto_adm
                if es_nuevo_formato and product_id:
                    producto_db = resolver_producto_adm(item_id=product_id, sku=sku)
                    if not producto_db:
                        errores.append({
                            "fila": row_idx,
                            "error": f"Product ID '{product_id}' no encontrado en la base de datos"
                        })
                        continue
                    item_id = product_id
                else:
                    producto_db = resolver_producto_adm(sku=sku)
                    if not producto_db:
                        errores.append({
                            "fila": row_idx,
                            "error": f"SKU '{sku}' no encontrado en la base de datos"
                        })
                        continue
                    item_id = producto_db.item_id
                
                # ✅ Validar ubicación física existe (permite agregar nuevas ubicaciones)
                ubicacion_fisica = UbicacionFisica.query.filter_by(
                    codigo=ubicacion,
                    activa=True
                ).first()
                
                if not ubicacion_fisica:
                    errores.append({
                        "fila": row_idx,
                        "error": f"Ubicación física '{ubicacion}' no existe o está inactiva. Debe crearse primero en el sistema de ubicaciones físicas."
                    })
                    continue
                
                # Agregar a la lista de ajustes válidos
                ajustes.append({
                    "sku": sku,
                    "ubicacion": ubicacion,
                    "cantidad": cantidad,
                    "notas": notas_fila or notas_generales or f"Ajuste masivo desde Excel - Fila {row_idx}",
                    "item_id": item_id,
                    "fila": row_idx
                })
                
            except Exception as e:
                errores.append({
                    "fila": row_idx,
                    "error": f"Error al procesar fila: {str(e)}"
                })
                continue
        
        if len(ajustes) == 0:
            return jsonify({
                "success": False,
                "error": "No se encontraron ajustes válidos en el archivo",
                "errores": errores
            }), 400
        
        # OPTIMIZACIÓN 1: Agrupar ajustes duplicados por (product_id, ubicacion)
        # Esto reduce el número de operaciones si hay filas duplicadas en el Excel
        ajustes_agrupados = {}
        for ajuste in ajustes:
            key = (ajuste['item_id'], ajuste['ubicacion'])
            if key not in ajustes_agrupados:
                ajustes_agrupados[key] = {
                    'sku': ajuste['sku'],
                    'ubicacion': ajuste['ubicacion'],
                    'cantidad': 0.0,
                    'notas': ajuste['notas'],  # Usar notas de la primera ocurrencia
                    'item_id': ajuste['item_id'],
                    'filas': []  # Para tracking
                }
            ajustes_agrupados[key]['cantidad'] += ajuste['cantidad']
            ajustes_agrupados[key]['filas'].append(ajuste['fila'])
        
        logger.info(f"Ajustes agrupados: {len(ajustes)} filas → {len(ajustes_agrupados)} combinaciones únicas")
        
        # OPTIMIZACIÓN 2: Cargar todos los stocks existentes de una vez (bulk load)
        # Esto reduce de N consultas a 1 consulta
        product_ids = list(set([a['item_id'] for a in ajustes_agrupados.values()]))
        ubicaciones = list(set([a['ubicacion'] for a in ajustes_agrupados.values()]))
        
        # ✅ MEJORADO: Cargar stocks con manejo de errores y fallback
        try:
            stocks_existentes = {
                (s.product_id, s.ubicacion): s 
                for s in StockUbicacion.query.filter(
                    StockUbicacion.product_id.in_(product_ids),
                    StockUbicacion.ubicacion.in_(ubicaciones)
                ).all()
            }
            logger.info(f"Stocks existentes cargados: {len(stocks_existentes)} registros")
        except Exception as e:
            logger.error(f"Error al cargar stocks existentes: {e}", exc_info=True)
            # Fallback: cargar stocks de forma más conservadora
            stocks_existentes = {}
            for product_id in product_ids[:100]:  # Limitar a primeros 100 para evitar timeout
                for ubicacion in ubicaciones:
                    stock = StockUbicacion.query.filter_by(
                        product_id=product_id,
                        ubicacion=ubicacion
                    ).first()
                    if stock:
                        stocks_existentes[(product_id, ubicacion)] = stock
            logger.info(f"Stocks existentes cargados (fallback): {len(stocks_existentes)} registros")
        
        # OPTIMIZACIÓN 3: Procesar en lotes para evitar timeout
        # ✅ REDUCIDO: De 500 a 250 para reducir carga en conexiones
        BATCH_SIZE = 250
        ajustes_lista = list(ajustes_agrupados.values())
        movimientos_creados = []
        productos_procesados = set()
        total_procesados = 0
        
        # ✅ MEJORADO: Sin no_autoflush, con flush explícito y limpieza de session
        for batch_start in range(0, len(ajustes_lista), BATCH_SIZE):
            batch = ajustes_lista[batch_start:batch_start + BATCH_SIZE]
            batch_num = batch_start // BATCH_SIZE + 1
            logger.info(f"Procesando lote {batch_num}: {len(batch)} ajustes")
            
            try:
                for ajuste in batch:
                    sku = ajuste['sku']
                    ubicacion = ajuste['ubicacion']
                    cantidad_nueva = ajuste['cantidad']
                    notas = ajuste['notas']
                    item_id = ajuste['item_id']
                    
                    # OPTIMIZACIÓN 4: Buscar stock usando product_id (corrige error de duplicado)
                    # Buscar en diccionario en memoria (O(1)) en lugar de consulta BD
                    key_stock = (item_id, ubicacion)
                    stock_ubic = stocks_existentes.get(key_stock)
                    
                    cantidad_anterior = 0.0
                    if stock_ubic:
                        cantidad_anterior = float(stock_ubic.cantidad)
                        stock_ubic.cantidad = cantidad_nueva
                        stock_ubic.updated_at = datetime.utcnow()
                        # Permite cantidad 0 para dejar ubicación vacía (ej: mover stock a otra)
                    else:
                        # Sin registro existente: solo crear si cantidad > 0 (evitar registros 0 innecesarios)
                        if cantidad_nueva == 0:
                            productos_procesados.add(sku)
                            total_procesados += 1
                            continue
                        stock_ubic = StockUbicacion(
                            product_id=item_id,
                            sku=sku,
                            ubicacion=ubicacion,
                            cantidad=cantidad_nueva,
                            updated_at=datetime.utcnow()
                        )
                        db.session.add(stock_ubic)
                        stocks_existentes[key_stock] = stock_ubic
                    
                    # Calcular diferencia
                    diferencia = cantidad_nueva - cantidad_anterior
                    
                    # Crear movimiento solo si hay diferencia
                    if diferencia != 0:
                        notas_movimiento = notas
                        if diferencia > 0:
                            notas_movimiento = f"{notas}. Anterior: {cantidad_anterior}, Nuevo: {cantidad_nueva}"
                        
                        movimiento = Movimiento(
                            tipo="ADJUSTMENT",
                            product_id=item_id,
                            sku=sku,
                            ubicacion_origen=ubicacion if diferencia < 0 else None,
                            ubicacion_destino=ubicacion if diferencia > 0 else None,
                            cantidad=abs(diferencia),
                            usuario_id=session.get('user_id'),
                            timestamp=timestamp_ajuste,  # Mismo timestamp para agrupar
                            notas=notas_movimiento
                        )
                        db.session.add(movimiento)
                        movimientos_creados.append(movimiento.to_dict())
                    
                    productos_procesados.add(sku)
                    total_procesados += 1
                
                # ✅ MEJORADO: Flush explícito antes de commit para evitar inconsistencias
                db.session.flush()
                
                # Commit parcial cada lote para evitar timeout
                db.session.commit()
                
                # NO usar expunge_all(): los objetos en stocks_existentes se reutilizan en lotes
                # siguientes. Expungirlos provoca DetachedInstanceError al acceder stock_ubic.cantidad
                
                logger.info(f"Lote {batch_num} procesado exitosamente")
                
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error al procesar lote {batch_num}: {e}", exc_info=True)
                
                # ✅ NUEVO: Si es error de conexión, intentar reconectar
                error_str = str(e).lower()
                if 'packet sequence' in error_str or 'command out of sync' in error_str or 'mysql server has gone away' in error_str:
                    logger.warning(f"Error de conexión detectado en lote {batch_num}, intentando reconectar...")
                    try:
                        db.session.close()
                        db.engine.dispose()  # Descartar pool para forzar nuevas conexiones
                        # Consumir resultado para evitar "Command Out of Sync"
                        db.session.execute(db.text("SELECT 1")).scalar()
                        logger.info(f"Reconexión exitosa después de error en lote {batch_num}")
                    except Exception as reconnect_error:
                        logger.error(f"Error al reconectar después de fallo en lote {batch_num}: {reconnect_error}")
                
                raise
        
        # ✅ LOGGING: Registrar errores (mensajes amarillos) en el log del servidor
        if errores:
            logger.warning(f"[CARGA MASIVA EXCEL] Procesamiento completado con {len(errores)} error(es) de {total_filas_excel} filas totales")
            # Agrupar errores por tipo para el log
            errores_por_tipo = {}
            for error in errores:
                tipo_error = error.get('error', 'Error desconocido')
                if tipo_error not in errores_por_tipo:
                    errores_por_tipo[tipo_error] = 0
                errores_por_tipo[tipo_error] += 1
            
            logger.warning(f"[CARGA MASIVA EXCEL] Resumen de errores: {errores_por_tipo}")
            # Mostrar primeros 10 errores como ejemplo
            if len(errores) <= 10:
                logger.warning(f"[CARGA MASIVA EXCEL] Errores detallados: {errores}")
            else:
                logger.warning(f"[CARGA MASIVA EXCEL] Primeros 10 errores: {errores[:10]}")
                logger.warning(f"[CARGA MASIVA EXCEL] ... y {len(errores) - 10} errores más")
        else:
            logger.info(f"[CARGA MASIVA EXCEL] Procesamiento exitoso sin errores: {len(ajustes)} ajustes, {len(movimientos_creados)} movimientos")

        # Actualizar/resolver discrepancias pendientes para productos afectados
        discrepancias_resueltas = actualizar_discrepancias_por_skus(productos_procesados)
        if discrepancias_resueltas > 0:
            logger.info(f"[CARGA MASIVA EXCEL] {discrepancias_resueltas} discrepancia(s) marcada(s) como resuelta(s)")
        
        return jsonify({
            "success": True,
            "message": f"Ajustes masivos procesados exitosamente",
            "total_ajustes": len(ajustes),
            "total_movimientos": len(movimientos_creados),
            "productos_unicos": len(productos_procesados),
            "discrepancias_resueltas": discrepancias_resueltas,
            "errores": errores if errores else [],
            "total_errores": len(errores)
        })
            
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al procesar archivo Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "Error al procesar archivo Excel",
            "message": str(e)
        }), 500


@ajustes_bp.route('/api/ajustes/<ajuste_id>/revertir', methods=['POST'])
@require_admin
def revertir_ajuste(ajuste_id):
    """Reverte un ajuste procesado (solo administradores) - Elimina movimientos y revierte stock"""
    try:
        from utils.helpers import parsear_ajuste_id
        timestamp, ubicacion = parsear_ajuste_id(ajuste_id)

        logger.info(f"[REVERTIR AJUSTE] ID recibido: {ajuste_id!r}, timestamp={timestamp}, ubicacion={ubicacion}")

        if not timestamp or not ubicacion:
            return jsonify({
                "success": False,
                "error": "ID de ajuste inválido"
            }), 400
        
        # Buscar movimientos por ubicacion_destino O ubicacion_origen (cubre ajustes positivos y negativos)
        movimientos = Movimiento.query.filter(
            Movimiento.tipo == 'ADJUSTMENT',
            Movimiento.timestamp == timestamp,
            or_(
                Movimiento.ubicacion_destino == ubicacion,
                Movimiento.ubicacion_origen == ubicacion
            )
        ).all()
        
        if not movimientos:
            return jsonify({
                "success": False,
                "error": "No se encontraron movimientos para este ajuste"
            }), 404
        
        # Determinar si es ubicación física o ADM
        es_ubicacion_fisica = False
        if ubicacion:
            ubicacion_fisica = UbicacionFisica.query.filter_by(
                codigo=ubicacion.upper(),
                activa=True
            ).first()
            es_ubicacion_fisica = ubicacion_fisica is not None
            if not es_ubicacion_fisica and len(ubicacion) >= 6 and any(c.isdigit() for c in ubicacion):
                es_ubicacion_fisica = True
        
        # Revertir stock y eliminar movimientos
        stock_revertido = 0
        cache_revertido = 0
        for movimiento in movimientos:
            notas = movimiento.notas or ''
            cantidad_anterior = None

            if 'Anterior:' in notas and 'Nuevo:' in notas:
                try:
                    partes_notas = notas.split('Anterior:')[1].split(',')
                    if partes_notas:
                        anterior_str = partes_notas[0].strip()
                        cantidad_anterior = float(anterior_str)
                except:
                    pass

            if es_ubicacion_fisica:
                # Revertir StockUbicacion (la ubicación real puede estar en destino u origen)
                ubicacion_mov = movimiento.ubicacion_destino or movimiento.ubicacion_origen
                if not ubicacion_mov:
                    ubicacion_mov = ubicacion

                if cantidad_anterior is not None:
                    stock_ubic = StockUbicacion.query.filter_by(
                        sku=movimiento.sku,
                        ubicacion=ubicacion_mov
                    ).first()
                    
                    if stock_ubic:
                        stock_ubic.cantidad = cantidad_anterior
                        stock_ubic.updated_at = datetime.utcnow()
                        stock_revertido += 1
                    else:
                        stock_ubic = StockUbicacion(
                            product_id=movimiento.product_id,
                            sku=movimiento.sku,
                            ubicacion=ubicacion_mov,
                            cantidad=cantidad_anterior,
                            updated_at=datetime.utcnow()
                        )
                        db.session.add(stock_ubic)
                        stock_revertido += 1
                else:
                    stock_ubic = StockUbicacion.query.filter_by(
                        sku=movimiento.sku,
                        ubicacion=ubicacion_mov
                    ).first()
                    if stock_ubic:
                        nueva_cantidad = max(0, float(stock_ubic.cantidad) - float(movimiento.cantidad))
                        stock_ubic.cantidad = nueva_cantidad
                        stock_ubic.updated_at = datetime.utcnow()
                        stock_revertido += 1
            else:
                # Revertir StockProductoADM (simétrico al registro)
                from utils.helpers import actualizar_cache_adm, eliminar_fila_cache_adm, resolver_producto_adm
                producto_db = resolver_producto_adm(item_id=movimiento.product_id, sku=movimiento.sku)
                ubicacion_adm_name = movimiento.ubicacion_destino or movimiento.ubicacion_origen or ubicacion
                loc_status = SyncLocationStatus.query.filter_by(location_name=ubicacion_adm_name).first()
                if not loc_status:
                    loc_status = SyncLocationStatus.query.filter(
                        SyncLocationStatus.location_name.ilike(f'%{ubicacion_adm_name}%')
                    ).first()
                loc_id = loc_status.location_id if loc_status else None

                if producto_db and loc_id:
                    fila_fue_creada = '[FILA_CREADA]' in notas
                    if fila_fue_creada:
                        eliminada = eliminar_fila_cache_adm(producto_db.id, loc_id)
                        if eliminada:
                            cache_revertido += 1
                            logger.info(f"[REVERTIR AJUSTE ADM] Fila creada eliminada: SKU={movimiento.sku}, ubicación={ubicacion_adm_name}")
                    elif cantidad_anterior is not None:
                        actualizar_cache_adm(
                            producto_id=producto_db.id,
                            location_id=loc_id,
                            valor_absoluto=cantidad_anterior,
                            location_name=ubicacion_adm_name
                        )
                        cache_revertido += 1
                        logger.info(f"[REVERTIR AJUSTE ADM] Cache restaurada: SKU={movimiento.sku}, valor={cantidad_anterior}")
                    else:
                        actualizar_cache_adm(
                            producto_id=producto_db.id,
                            location_id=loc_id,
                            delta=-float(movimiento.cantidad),
                            location_name=ubicacion_adm_name
                        )
                        cache_revertido += 1

            db.session.delete(movimiento)
        
        db.session.commit()
        
        mensaje = f"Ajuste revertido exitosamente. Se eliminaron {len(movimientos)} movimiento(s)."
        if es_ubicacion_fisica:
            mensaje += f" Se revirtió el stock en {stock_revertido} ubicación(es) física(s)."
        else:
            mensaje += f" Se revirtió la cache ADM en {cache_revertido} registro(s)."
        
        return jsonify({
            "success": True,
            "message": mensaje,
            "movimientos_eliminados": len(movimientos),
            "stock_revertido": stock_revertido if es_ubicacion_fisica else 0,
            "cache_revertido": cache_revertido
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al revertir ajuste: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "Error al revertir ajuste",
            "message": str(e)
        }), 500

