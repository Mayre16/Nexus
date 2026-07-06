"""
Abastecimiento: políticas min/max por ubicación ADM y consultas para reposición (fase 1).
Solo administradores (require_admin).
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from flask import Blueprint, Response, current_app, jsonify, request
from sqlalchemy import or_, select

from database import db
from database.models import AbastecimientoPolitica, ProductoADM, StockProductoADM, SyncLocationStatus
from routes.auth import require_admin
from utils.helpers import obtener_mapa_stock_vigente, obtener_stock_vigente

import logging

logger = logging.getLogger(__name__)

abastecimiento_bp = Blueprint('abastecimiento', __name__)

_ERR_OPENPYXL = (
    'Falta la librería openpyxl para generar o leer Excel (.xlsx). '
    'Ejecute en el entorno del proyecto: pip install openpyxl '
    'o pip install -r requirements.txt'
)


def _openpyxl_classes():
    """Retorna (Workbook, load_workbook) o (None, None) si no está instalado."""
    try:
        from openpyxl import Workbook, load_workbook
        return Workbook, load_workbook
    except ModuleNotFoundError:
        return None, None


# Etiquetas en Excel (columna Estado). La hoja Leyenda documenta los mismos valores.
_ESTADO_EXCEL_ETIQUETA = {
    'sin_config': 'Sin configuración',
    'bajo_minimo': 'Bajo mínimo',
    'en_rango': 'En rango',
    'sobre_maximo': 'Sobre máximo',
    'inactivo': 'Política inactiva',
}

# Encabezados estándar export / import (misma convención que Ajustes: Product ID = item_id ADM)
ABASTECIMIENTO_EXCEL_HEADERS = [
    'Product ID',
    'SKU',
    'Nombre',
    'Stock actual',
    'Mínimo',
    'Máximo',
    'Sugerido',
    'Estado',
    'Prioridad',
    'Incluido',
]


def _estado_para_excel(estado_code: str) -> str:
    return _ESTADO_EXCEL_ETIQUETA.get(estado_code, estado_code)


def _prioridad_excel_si_no(pol: AbastecimientoPolitica | None) -> str:
    """Prioridad en Excel = política activa (SI/NO), no el nivel calculado."""
    if pol and pol.activo:
        return 'SI'
    return 'NO'


def _incluido_excel_si_no(pol: AbastecimientoPolitica | None) -> str:
    if pol and pol.es_base_abastecimiento:
        return 'SI'
    return 'NO'


def _sugerido_para_excel(sug: float | None):
    """Valor numérico para celda Sugerido; vacío si no aplica."""
    if sug is None:
        return ''
    return float(sug)


def _producto_id_desde_celda(val) -> int | None:
    """Solo id numérico local (compatibilidad con Excels antiguos)."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return val if val > 0 else None
    if isinstance(val, float):
        if val != val or val < 0:  # nan
            return None
        i = int(val)
        return i if float(i) == val and i > 0 else None
    s = str(val).strip()
    if not s:
        return None
    try:
        d = Decimal(s.replace(',', '.'))
        i = int(d)
        if d != i or i <= 0:
            return None
        return i
    except Exception:
        return None


def _producto_adm_desde_identificador_excel(val) -> ProductoADM | None:
    """
    Igual criterio que Ajustes (catálogo Excel): primero item_id (GUID ADM),
    luego id entero local por compatibilidad con archivos viejos.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    p = ProductoADM.query.filter_by(item_id=s).first()
    if p:
        return p
    pid = _producto_id_desde_celda(val)
    if pid is not None:
        return db.session.get(ProductoADM, pid)
    return None


def _politica_activa_desde_celda(val) -> bool:
    """Celda Prioridad/Activo: SI/NO, vacío = activa por defecto."""
    if val is None:
        return True
    s = str(val).strip()
    if not s:
        return True
    a = s.lower()
    if a in ('no', 'n', '0', 'false', 'f', 'inactiva', 'inactivo', 'off'):
        return False
    return True


def _incluido_desde_celda(val):
    """
    Columna Incluido/Base:
      - SI/NO => True/False
      - vacío  => None (no modificar el valor actual)
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    a = s.lower()
    if a in ('si', 'sí', 's', '1', 'true', 't', 'on', 'incluido', 'base'):
        return True
    if a in ('no', 'n', '0', 'false', 'f', 'off', 'no incluido', 'excluir'):
        return False
    return None


def _rellenar_hoja_leyenda(ws) -> None:
    ws.append(['Abastecimiento — Leyenda'])
    ws.append([])
    ws.append(['Columna SUGERIDO (solo informativa en el export; al importar se ignora).'])
    ws.append([
        'Cantidad sugerida para reposición hasta el máximo cuando el stock está bajo el mínimo y la política está activa. '
        'Si no aplica, va vacío.',
    ])
    ws.append([])
    ws.append(['Columna ESTADO (solo informativa en el export; al importar se ignora).'])
    ws.append(['Valores que puede mostrar el sistema:'])
    for code, label in _ESTADO_EXCEL_ETIQUETA.items():
        ws.append([label, f'código: {code}'])
    ws.append([])
    ws.append(['Columna PRIORIDAD (sí se importa): indica si la política de abastecimiento está activa.'])
    ws.append(['SI', 'Activa: se usan mínimo/máximo y sugerencias.'])
    ws.append(['NO', 'Inactiva: no se sugieren envíos aunque existan mín/máx.'])
    ws.append([])
    ws.append(['Columna INCLUIDO (sí se importa): si el producto entra al universo base de abastecimiento.'])
    ws.append(['SI', 'Incluido en la vista principal (universo incluidos).'])
    ws.append(['NO', 'No incluido en la vista principal (visible por filtros).'])
    ws.append([])
    ws.append([
        'Importación: la clave es Product ID = item_id de ADM Cloud (mismo que el catálogo de Ajustes). '
        'También se acepta el id numérico interno del WMS en archivos antiguos. '
        'Se actualizan Mínimo, Máximo, Prioridad (SI/NO) e Incluido (SI/NO). '
        'Sugerido y Estado del archivo, si vienen, se ignoran al importar.',
    ])


def _resolver_location_id() -> tuple[str | None, str | None, str | None]:
    """
    Devuelve (location_id, location_name_display, error_message).
    """
    lid = (current_app.config.get('ABASTECIMIENTO_LOCATION_ID') or '').strip()
    name_cfg = (current_app.config.get('ABASTECIMIENTO_LOCATION_NAME') or 'Mirador Sur').strip()

    if lid:
        row = SyncLocationStatus.query.filter_by(location_id=lid).first()
        name = row.location_name if row else name_cfg
        return lid, name, None

    row = SyncLocationStatus.query.filter(
        SyncLocationStatus.location_name.ilike(f'%{name_cfg}%')
    ).first()
    if row:
        return row.location_id, row.location_name, None
    return None, None, (
        f'No se encontró la ubicación "{name_cfg}" en sync_locations_status. '
        'Configure ABASTECIMIENTO_LOCATION_ID o sincronice ubicaciones.'
    )


# Mientras sync está "running" en la ubicación de abastecimiento, no ejecutar consultas pesadas.
ABAST_ERROR_SYNC_EN_CURSO = 'sync_en_curso'
ABAST_MSG_SYNC_EN_CURSO = (
    'Lo sentimos, debes esperar unos minutos: está corriendo una sincronización. '
    'Recarga la página hasta que ya no aparezca este mensaje.'
)


def _json_sync_en_curso():
    return jsonify({
        'success': False,
        'error': ABAST_ERROR_SYNC_EN_CURSO,
        'message': ABAST_MSG_SYNC_EN_CURSO,
    }), 503


def _sync_bloquea_abastecimiento(location_id: str | None) -> bool:
    if not location_id:
        return False
    row = SyncLocationStatus.query.filter_by(location_id=location_id).first()
    return bool(row and row.status == 'running')


def _normalizar_universo(v: str | None) -> str:
    x = (v or 'incluidos').strip().lower()
    if x not in ('incluidos', 'no_incluidos', 'todos'):
        return 'incluidos'
    return x


def _subquery_ids_incluidos(location_id: str):
    return (
        select(AbastecimientoPolitica.producto_id)
        .where(
            AbastecimientoPolitica.location_id == location_id,
            AbastecimientoPolitica.es_base_abastecimiento.is_(True),
        )
        .scalar_subquery()
    )


def _stock_actual(producto_id: int, location_id: str) -> tuple[float, datetime | None]:
    sp = obtener_stock_vigente(producto_id, location_id)
    if not sp:
        return 0.0, None
    st = float(sp.stock) if sp.stock is not None else 0.0
    ts = sp.updated_at if hasattr(sp, 'updated_at') and sp.updated_at else None
    return st, ts


def _producto_ids_stock_positivo_en_ubicacion(location_id: str) -> set[int]:
    """Productos con stock > 0 en cache ADM para la ubicación (run vigente o legacy)."""
    estado_sync = SyncLocationStatus.query.filter_by(location_id=location_id).first()
    if estado_sync and estado_sync.current_run_id:
        return set(
            r[0] for r in db.session.query(StockProductoADM.producto_id).filter(
                StockProductoADM.location_id == location_id,
                StockProductoADM.sync_run_id == estado_sync.current_run_id,
                StockProductoADM.stock > 0,
            ).all()
        )
    return set(
        r[0] for r in db.session.query(StockProductoADM.producto_id).filter(
            StockProductoADM.location_id == location_id,
            StockProductoADM.sync_run_id.is_(None),
            StockProductoADM.stock > 0,
        ).all()
    )


def _ids_por_prioridad(location_id: str, prioridad_code: str) -> set[int]:
    """Conjunto de producto_id cuya prioridad calculada coincide (catálogo activo)."""
    pols = {
        p.producto_id: p
        for p in AbastecimientoPolitica.query.filter_by(location_id=location_id).all()
    }
    pids = [r[0] for r in db.session.query(ProductoADM.id).filter(ProductoADM.activo == True).all()]
    stock_map = obtener_mapa_stock_vigente(pids, location_id)
    out: set[int] = set()
    for pid in pids:
        pol = pols.get(pid)
        stock = stock_map.get(pid, (0.0, None))[0]
        tiene = pol is not None
        pa = bool(pol and pol.activo)
        smin = float(pol.stock_min) if pol else 0.0
        smax = float(pol.stock_max) if pol else 0.0
        pr = _prioridad(stock, smin, smax, tiene, pa)
        if pr == prioridad_code:
            out.add(pid)
    return out


def _prioridad(stock: float, stock_min: float, stock_max: float, tiene_config: bool, politica_activa: bool) -> str:
    if not tiene_config or not politica_activa:
        return 'sin_config'
    if stock_min <= 0:
        return 'normal'
    if stock <= 0:
        return 'critica'
    if stock < stock_min * 0.5:
        return 'alta'
    if stock < stock_min:
        return 'media'
    if stock > stock_max:
        return 'sobre_maximo'
    return 'normal'


def _estado_visual(stock: float, stock_min: float, stock_max: float, tiene_config: bool, politica_activa: bool) -> str:
    if not tiene_config:
        return 'sin_config'
    if not politica_activa:
        return 'inactivo'
    if stock < stock_min:
        return 'bajo_minimo'
    if stock > stock_max:
        return 'sobre_maximo'
    return 'en_rango'


def _sugerido_enviar(stock: float, stock_min: float, stock_max: float, tiene_config: bool, politica_activa: bool) -> float | None:
    if not tiene_config or not politica_activa:
        return None
    if stock >= stock_min:
        return 0.0
    return max(0.0, round(stock_max - stock, 2))


def _serialize_row(
    p: ProductoADM,
    location_id: str,
    pol: AbastecimientoPolitica | None,
    stock_cached: tuple[float, datetime | None] | None = None,
) -> dict:
    if stock_cached is not None:
        stock, ts_stock = stock_cached
    else:
        stock, ts_stock = _stock_actual(p.id, location_id)
    tiene = pol is not None
    pa = bool(pol and pol.activo)
    smin = float(pol.stock_min) if pol else 0.0
    smax = float(pol.stock_max) if pol else 0.0
    est = _estado_visual(stock, smin, smax, tiene, pa)
    sug = _sugerido_enviar(stock, smin, smax, tiene, pa)
    pri = _prioridad(stock, smin, smax, tiene, pa)
    return {
        'producto_id': p.id,
        'sku': p.sku,
        'nombre': p.nombre or '',
        'stock_actual': stock,
        'stock_actual_updated_at': ts_stock.isoformat() + 'Z' if ts_stock else None,
        'stock_min': smin,
        'stock_max': smax,
        'activo_politica': pa,
        'es_base_abastecimiento': bool(pol and pol.es_base_abastecimiento),
        'tiene_configuracion': tiene,
        'estado': est,
        'prioridad': pri,
        'cantidad_sugerida': sug,
        'observacion': pol.observacion if pol else None,
        'politica_updated_at': pol.updated_at.isoformat() + 'Z' if pol and pol.updated_at else None,
    }


@abastecimiento_bp.route('/abastecimiento')
@require_admin
def abastecimiento_page():
    from flask import render_template
    return render_template('abastecimiento.html')


@abastecimiento_bp.route('/api/abastecimiento/meta', methods=['GET'])
@require_admin
def api_meta():
    lid, lname, err = _resolver_location_id()
    sync = SyncLocationStatus.query.filter_by(location_id=lid).first() if lid else None
    bloqueo = _sync_bloquea_abastecimiento(lid) if lid else False
    return jsonify({
        'success': True,
        'location_id': lid,
        'location_name': lname,
        'error_config': err,
        'last_sync_at': (sync.last_sync_at.isoformat() + 'Z') if sync and sync.last_sync_at else None,
        'sync_status': sync.status if sync else None,
        'sync_bloqueando': bloqueo,
        'sync_bloqueo_mensaje': ABAST_MSG_SYNC_EN_CURSO if bloqueo else None,
    })


@abastecimiento_bp.route('/api/abastecimiento/kpis', methods=['GET'])
@require_admin
def api_kpis():
    lid, _, err = _resolver_location_id()
    if err or not lid:
        return jsonify({'success': False, 'error': err or 'Ubicación no configurada'}), 400
    if _sync_bloquea_abastecimiento(lid):
        return _json_sync_en_curso()

    universo = _normalizar_universo(request.args.get('universo'))
    incl_sub = _subquery_ids_incluidos(lid)

    activos_q = ProductoADM.query.filter(ProductoADM.activo == True)
    if universo == 'incluidos':
        activos_q = activos_q.filter(ProductoADM.id.in_(incl_sub))
    elif universo == 'no_incluidos':
        activos_q = activos_q.filter(~ProductoADM.id.in_(incl_sub))
    activos_ids = [r[0] for r in activos_q.with_entities(ProductoADM.id).all()]
    total_activos = len(activos_ids)

    if activos_ids:
        con_pol = (
            AbastecimientoPolitica.query.filter(
                AbastecimientoPolitica.location_id == lid,
                AbastecimientoPolitica.producto_id.in_(activos_ids),
            ).count()
        )
    else:
        con_pol = 0
    sin_pol = max(0, total_activos - con_pol)

    politicas_q = AbastecimientoPolitica.query.filter_by(location_id=lid, activo=True)
    if activos_ids:
        politicas_q = politicas_q.filter(AbastecimientoPolitica.producto_id.in_(activos_ids))
    else:
        politicas_q = politicas_q.filter(False)
    politicas = politicas_q.all()
    pids_pol = [pol.producto_id for pol in politicas]
    stock_map_pol = obtener_mapa_stock_vigente(pids_pol, lid)
    bajo = 0
    sobre = 0
    suma_sugerida = 0.0
    for pol in politicas:
        st = stock_map_pol.get(pol.producto_id, (0.0, None))[0]
        smin = float(pol.stock_min)
        smax = float(pol.stock_max)
        if st < smin:
            bajo += 1
            suma_sugerida += max(0.0, smax - st)
        if st > smax:
            sobre += 1

    pos_ids = _producto_ids_stock_positivo_en_ubicacion(lid)
    if activos_ids:
        pos_ids = set(pid for pid in pos_ids if pid in set(activos_ids))
    else:
        pos_ids = set()
    if pos_ids:
        cero = (
            db.session.query(ProductoADM.id)
            .filter(ProductoADM.activo == True, ~ProductoADM.id.in_(pos_ids))
            .count()
        )
    else:
        cero = total_activos

    pct_rango = None
    if con_pol > 0:
        en_rango = con_pol - bajo - sobre
        pct_rango = round(100.0 * max(0, en_rango) / con_pol, 1)

    return jsonify({
        'success': True,
        'universo': universo,
        'total_productos_activos': total_activos,
        'con_politica': con_pol,
        'sin_politica': sin_pol,
        'bajo_minimo': bajo,
        'sobre_maximo': sobre,
        'stock_cero': cero,
        'suma_cantidad_sugerida': round(suma_sugerida, 2),
        'porcentaje_en_rango': pct_rango,
    })


@abastecimiento_bp.route('/api/abastecimiento/productos', methods=['GET'])
@require_admin
def api_productos():
    lid, _, err = _resolver_location_id()
    if err or not lid:
        return jsonify({'success': False, 'error': err or 'Ubicación no configurada'}), 400
    if _sync_bloquea_abastecimiento(lid):
        return _json_sync_en_curso()

    page = max(1, request.args.get('page', 1, type=int)
        or 1)
    per_page = min(100, max(10, request.args.get('per_page', 50, type=int) or 50))
    q = (request.args.get('q') or '').strip()
    estado = (request.args.get('estado') or 'todos').strip().lower()
    sf_raw = request.args.get('stock_filtro')
    if sf_raw is not None and str(sf_raw).strip() != '':
        stock_filtro = str(sf_raw).strip().lower()
        if stock_filtro not in ('todos', 'solo_cero', 'solo_positivo'):
            stock_filtro = 'todos'
    elif request.args.get('solo_stock_cero', '').lower() in ('1', 'true', 'si', 'sí'):
        stock_filtro = 'solo_cero'
    else:
        stock_filtro = 'todos'
    prioridad_f = (request.args.get('prioridad') or 'todos').strip().lower()
    _prios_ok = ('todos', 'sin_config', 'critica', 'alta', 'media', 'normal', 'sobre_maximo')
    if prioridad_f not in _prios_ok:
        prioridad_f = 'todos'
    universo = _normalizar_universo(request.args.get('universo'))

    base = ProductoADM.query.filter(ProductoADM.activo == True)
    incl_sub = _subquery_ids_incluidos(lid)
    if universo == 'incluidos':
        base = base.filter(ProductoADM.id.in_(incl_sub))
    elif universo == 'no_incluidos':
        base = base.filter(~ProductoADM.id.in_(incl_sub))
    if q:
        term = f'%{q}%'
        base = base.filter(or_(ProductoADM.sku.ilike(term), ProductoADM.nombre.ilike(term)))

    if estado == 'sin_config':
        sub = (
            select(AbastecimientoPolitica.producto_id)
            .where(AbastecimientoPolitica.location_id == lid)
            .scalar_subquery()
        )
        base = base.filter(~ProductoADM.id.in_(sub))

    elif estado == 'bajo_minimo':
        pols_bm = AbastecimientoPolitica.query.filter_by(location_id=lid, activo=True).all()
        pids_bm = [pol.producto_id for pol in pols_bm]
        sm_bm = obtener_mapa_stock_vigente(pids_bm, lid)
        ids = [
            pol.producto_id for pol in pols_bm
            if sm_bm.get(pol.producto_id, (0.0, None))[0] < float(pol.stock_min)
        ]
        if not ids:
            return jsonify({
                'success': True,
                'items': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'pages': 0,
            })
        base = base.filter(ProductoADM.id.in_(ids))

    elif estado == 'en_rango':
        pols_er = AbastecimientoPolitica.query.filter_by(location_id=lid, activo=True).all()
        pids_er = [pol.producto_id for pol in pols_er]
        sm_er = obtener_mapa_stock_vigente(pids_er, lid)
        ids = []
        for pol in pols_er:
            st = sm_er.get(pol.producto_id, (0.0, None))[0]
            smin = float(pol.stock_min)
            smax = float(pol.stock_max)
            if smin <= st <= smax:
                ids.append(pol.producto_id)
        if not ids:
            return jsonify({
                'success': True,
                'items': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'pages': 0,
            })
        base = base.filter(ProductoADM.id.in_(ids))

    elif estado == 'sobre_maximo':
        pols_sm = AbastecimientoPolitica.query.filter_by(location_id=lid, activo=True).all()
        pids_sm = [pol.producto_id for pol in pols_sm]
        sm_sm = obtener_mapa_stock_vigente(pids_sm, lid)
        ids = [
            pol.producto_id for pol in pols_sm
            if sm_sm.get(pol.producto_id, (0.0, None))[0] > float(pol.stock_max)
        ]
        if not ids:
            return jsonify({
                'success': True,
                'items': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'pages': 0,
            })
        base = base.filter(ProductoADM.id.in_(ids))

    if stock_filtro == 'solo_cero':
        pos_ids = _producto_ids_stock_positivo_en_ubicacion(lid)
        if pos_ids:
            base = base.filter(~ProductoADM.id.in_(pos_ids))
    elif stock_filtro == 'solo_positivo':
        pos_ids = _producto_ids_stock_positivo_en_ubicacion(lid)
        if not pos_ids:
            return jsonify({
                'success': True,
                'items': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'pages': 0,
            })
        base = base.filter(ProductoADM.id.in_(pos_ids))

    if prioridad_f != 'todos':
        prio_ids = _ids_por_prioridad(lid, prioridad_f)
        if not prio_ids:
            return jsonify({
                'success': True,
                'items': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'pages': 0,
            })
        base = base.filter(ProductoADM.id.in_(prio_ids))

    pag = base.order_by(ProductoADM.sku.asc()).paginate(page=page, per_page=per_page, error_out=False)
    pids = [p.id for p in pag.items]
    pol_map = {}
    if pids:
        for pol in AbastecimientoPolitica.query.filter(
            AbastecimientoPolitica.location_id == lid,
            AbastecimientoPolitica.producto_id.in_(pids),
        ).all():
            pol_map[pol.producto_id] = pol

    stock_page = obtener_mapa_stock_vigente(pids, lid)
    items = [
        _serialize_row(p, lid, pol_map.get(p.id), stock_page.get(p.id))
        for p in pag.items
    ]

    return jsonify({
        'success': True,
        'universo': universo,
        'items': items,
        'total': pag.total,
        'page': pag.page,
        'per_page': pag.per_page,
        'pages': pag.pages,
    })


@abastecimiento_bp.route('/api/abastecimiento/politica', methods=['PUT'])
@require_admin
def api_put_politica():
    lid, _, err = _resolver_location_id()
    if err or not lid:
        return jsonify({'success': False, 'error': err or 'Ubicación no configurada'}), 400
    if _sync_bloquea_abastecimiento(lid):
        return _json_sync_en_curso()

    data = request.json or {}
    producto_id = data.get('producto_id')
    if not producto_id:
        return jsonify({'success': False, 'error': 'producto_id requerido'}), 400

    p = db.session.get(ProductoADM, int(producto_id))
    if not p:
        return jsonify({'success': False, 'error': 'Producto no encontrado'}), 404

    try:
        smin = Decimal(str(data.get('stock_min')))
        smax = Decimal(str(data.get('stock_max')))
    except (InvalidOperation, TypeError, ValueError):
        return jsonify({'success': False, 'error': 'stock_min y stock_max deben ser numéricos'}), 400

    if smin < 0 or smax < 0:
        return jsonify({'success': False, 'error': 'No se permiten valores negativos'}), 400
    if smin > smax:
        return jsonify({'success': False, 'error': 'stock_min no puede ser mayor que stock_max'}), 400

    activo = bool(data.get('activo', True))
    es_base = bool(data.get('es_base_abastecimiento', False))
    obs = (data.get('observacion') or '').strip()[:500] or None
    from flask import session
    uid = session.get('user_id')

    pol = AbastecimientoPolitica.query.filter_by(producto_id=p.id, location_id=lid).first()
    if not pol:
        pol = AbastecimientoPolitica(
            producto_id=p.id,
            location_id=lid,
            stock_min=smin,
            stock_max=smax,
            activo=activo,
            es_base_abastecimiento=es_base,
            observacion=obs,
            updated_by_user_id=uid,
        )
        db.session.add(pol)
    else:
        pol.stock_min = smin
        pol.stock_max = smax
        pol.activo = activo
        pol.es_base_abastecimiento = es_base
        pol.observacion = obs
        pol.updated_by_user_id = uid

    db.session.commit()
    return jsonify({'success': True, 'item': _serialize_row(p, lid, pol)})


@abastecimiento_bp.route('/api/abastecimiento/import/preview', methods=['POST'])
@require_admin
def api_import_preview():
    lid, _, err = _resolver_location_id()
    if err or not lid:
        return jsonify({'success': False, 'error': err or 'Ubicación no configurada'}), 400
    if _sync_bloquea_abastecimiento(lid):
        return _json_sync_en_curso()

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'error': 'Archivo requerido'}), 400

    rows, file_errors = _parse_import_file(f)
    if file_errors:
        return jsonify({'success': False, 'error': file_errors[0], 'errors': file_errors}), 400

    ok, bad = _validate_import_rows(rows, lid)
    return jsonify({
        'success': True,
        'filas_ok': len(ok),
        'filas_error': len(bad),
        'muestra_ok': [{
            'producto_id': x['producto_id'],
            'item_id': x.get('item_id'),
            'sku': x['sku'],
            'stock_min': float(x['stock_min']),
            'stock_max': float(x['stock_max']),
            'activo': x['activo'],
            'es_base_abastecimiento': x.get('es_base_abastecimiento'),
        } for x in ok[:50]],
        'errores': bad,
    })


@abastecimiento_bp.route('/api/abastecimiento/import/apply', methods=['POST'])
@require_admin
def api_import_apply():
    lid, _, err = _resolver_location_id()
    if err or not lid:
        return jsonify({'success': False, 'error': err or 'Ubicación no configurada'}), 400
    if _sync_bloquea_abastecimiento(lid):
        return _json_sync_en_curso()

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'error': 'Archivo requerido'}), 400

    rows, file_errors = _parse_import_file(f)
    if file_errors:
        return jsonify({'success': False, 'error': file_errors[0]}), 400

    ok, bad = _validate_import_rows(rows, lid)
    if bad:
        return jsonify({
            'success': False,
            'error': 'Hay filas con error. Corrija el archivo o use solo filas válidas.',
            'filas_error': len(bad),
            'errores': bad,
        }), 400

    from flask import session
    uid = session.get('user_id')

    aplicadas = 0
    for row in ok:
        p = db.session.get(ProductoADM, row['producto_id'])
        pol = AbastecimientoPolitica.query.filter_by(producto_id=p.id, location_id=lid).first()
        if not pol:
            pol = AbastecimientoPolitica(
                producto_id=p.id,
                location_id=lid,
                stock_min=row['stock_min'],
                stock_max=row['stock_max'],
                activo=row['activo'],
                es_base_abastecimiento=(
                    bool(row['es_base_abastecimiento']) if row.get('es_base_abastecimiento') is not None else False
                ),
                updated_by_user_id=uid,
            )
            db.session.add(pol)
        else:
            pol.stock_min = row['stock_min']
            pol.stock_max = row['stock_max']
            pol.activo = row['activo']
            if row.get('es_base_abastecimiento') is not None:
                pol.es_base_abastecimiento = bool(row['es_base_abastecimiento'])
            pol.updated_by_user_id = uid
        aplicadas += 1

    db.session.commit()
    return jsonify({'success': True, 'aplicadas': aplicadas})


def _parse_import_file(f):
    """Parsea Excel/CSV. Clave: Product ID (item_id ADM, como Ajustes). Actualiza min, max, activo e incluido."""
    name = (f.filename or '').lower()
    raw = f.read()
    try:
        if name.endswith('.xlsx'):
            _, load_workbook = _openpyxl_classes()
            if load_workbook is None:
                return [], [_ERR_OPENPYXL]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = list(ws.iter_rows(values_only=True))
        else:
            import csv
            text = raw.decode('utf-8-sig', errors='replace')
            r = csv.reader(io.StringIO(text))
            rows_iter = list(r)
    except Exception as e:
        return [], [f'No se pudo leer el archivo: {e}']

    if not rows_iter:
        return [], ['Archivo vacío']

    header = [str(c).strip().lower() if c is not None else '' for c in rows_iter[0]]
    col_pid = None
    col_min = None
    col_max = None
    col_activo = None
    col_incluido = None
    for i, h in enumerate(header):
        if h in (
            'product id', 'product_id', 'productid',
            'item_id', 'item id',
            'id producto', 'id_producto', 'producto_id', 'id del producto',
        ):
            col_pid = i
        if h in ('minimo', 'mínimo', 'min', 'stock_min', 'stock min'):
            col_min = i
        if h in ('maximo', 'máximo', 'max', 'stock_max', 'stock max'):
            col_max = i
        if h in (
            'prioridad', 'prioridad activa', 'politica activa', 'política activa',
            'activo', 'active', 'activa',
        ):
            if col_activo is None:
                col_activo = i
        if h in (
            'incluido', 'incluir', 'base', 'base abastecimiento',
            'es_base_abastecimiento', 'visible_principal', 'universo base',
        ):
            if col_incluido is None:
                col_incluido = i

    if col_pid is None:
        return [], [
            'Encabezado requerido: "Product ID" (item_id ADM, igual que en Ajustes). '
            'Descargue Excel bajo mínimo o completo desde esta pantalla.',
        ]

    if col_min is None or col_max is None:
        return [], ['Encabezados requeridos: Mínimo y Máximo (minimo/maximo).']

    out = []
    for idx, row in enumerate(rows_iter[1:], start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        raw = row[col_pid] if col_pid < len(row) else None
        raw_s = str(raw).strip() if raw is not None else ''
        prod = _producto_adm_desde_identificador_excel(raw)
        if prod is None:
            out.append({
                'fila': idx,
                'item_id': raw_s or None,
                'error': 'Product ID no encontrado en catálogo (revise GUID o use el Excel exportado)',
            })
            continue
        pid = prod.id
        try:
            vmin = row[col_min] if col_min < len(row) else None
            vmax = row[col_max] if col_max < len(row) else None
            vmin = Decimal(str(vmin).replace(',', '.'))
            vmax = Decimal(str(vmax).replace(',', '.'))
        except Exception:
            out.append({
                'fila': idx,
                'producto_id': pid,
                'item_id': prod.item_id,
                'error': 'Mínimo/Máximo no numéricos',
            })
            continue

        act = _politica_activa_desde_celda(
            row[col_activo] if col_activo is not None and col_activo < len(row) else None
        )
        inc = _incluido_desde_celda(
            row[col_incluido] if col_incluido is not None and col_incluido < len(row) else None
        )

        out.append({
            'fila': idx,
            'producto_id': pid,
            'item_id': prod.item_id,
            'stock_min': vmin,
            'stock_max': vmax,
            'activo': act,
            'es_base_abastecimiento': inc,
        })

    return out, []


def _validate_import_rows(rows: list, lid: str):
    """Valida filas parseadas por ID de producto."""
    ok = []
    bad = []
    for r in rows:
        if 'error' in r:
            bad.append(r)
            continue
        pid = r['producto_id']
        p = db.session.get(ProductoADM, pid)
        if not p:
            bad.append({
                'fila': r.get('fila'),
                'producto_id': pid,
                'item_id': r.get('item_id'),
                'error': 'Producto no existe en el catálogo',
            })
            continue
        if not p.activo:
            bad.append({
                'fila': r.get('fila'),
                'producto_id': pid,
                'item_id': p.item_id,
                'sku': p.sku,
                'error': 'Producto inactivo en catálogo',
            })
            continue
        smin = r['stock_min']
        smax = r['stock_max']
        if smin < 0 or smax < 0:
            bad.append({
                'fila': r.get('fila'),
                'producto_id': pid,
                'item_id': p.item_id,
                'sku': p.sku,
                'error': 'Valores negativos no permitidos',
            })
            continue
        if smin > smax:
            bad.append({
                'fila': r.get('fila'),
                'producto_id': pid,
                'item_id': p.item_id,
                'sku': p.sku,
                'error': 'Mínimo mayor que máximo',
            })
            continue
        ok.append({
            'producto_id': p.id,
            'item_id': p.item_id,
            'sku': p.sku,
            'stock_min': smin,
            'stock_max': smax,
            'activo': r['activo'],
            'es_base_abastecimiento': r.get('es_base_abastecimiento'),
        })
    return ok, bad


@abastecimiento_bp.route('/api/abastecimiento/export', methods=['GET'])
@require_admin
def export_excel():
    modo = (request.args.get('modo') or 'bajo_minimo').strip().lower()
    universo = _normalizar_universo(request.args.get('universo'))
    lid, lname, err = _resolver_location_id()
    if err or not lid:
        return jsonify({'success': False, 'error': err or 'Ubicación no configurada'}), 400
    if _sync_bloquea_abastecimiento(lid):
        return _json_sync_en_curso()

    Workbook, _ = _openpyxl_classes()
    if Workbook is None:
        return jsonify({'success': False, 'error': _ERR_OPENPYXL}), 503
    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos'
    ws.append(list(ABASTECIMIENTO_EXCEL_HEADERS))

    productos_q = ProductoADM.query.filter_by(activo=True)
    incl_sub = _subquery_ids_incluidos(lid)
    if universo == 'incluidos':
        productos_q = productos_q.filter(ProductoADM.id.in_(incl_sub))
    elif universo == 'no_incluidos':
        productos_q = productos_q.filter(~ProductoADM.id.in_(incl_sub))
    productos = productos_q.order_by(ProductoADM.sku.asc()).all()
    pol_all = {
        p.producto_id: p
        for p in AbastecimientoPolitica.query.filter_by(location_id=lid).all()
    }
    pids_export = [p.id for p in productos]
    stock_export = obtener_mapa_stock_vigente(pids_export, lid)

    for p in productos:
        pol = pol_all.get(p.id)
        row = _serialize_row(p, lid, pol, stock_export.get(p.id))
        if modo == 'bajo_minimo':
            if row['estado'] != 'bajo_minimo' or not row.get('tiene_configuracion'):
                continue
            if row.get('cantidad_sugerida') is None or row['cantidad_sugerida'] <= 0:
                continue

        ws.append([
            p.item_id,
            row['sku'],
            row['nombre'],
            row['stock_actual'],
            row['stock_min'],
            row['stock_max'],
            _sugerido_para_excel(row.get('cantidad_sugerida')),
            _estado_para_excel(row['estado']),
            _prioridad_excel_si_no(pol),
            _incluido_excel_si_no(pol),
        ])

    ws_ley = wb.create_sheet('Leyenda')
    _rellenar_hoja_leyenda(ws_ley)

    buf = io.BytesIO()
    wb.save(buf)
    suf = 'bajo_minimo' if modo == 'bajo_minimo' else 'completo'
    filename = f'abastecimiento_{suf}_{universo}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx'
    data = buf.getvalue()
    # send_file(BytesIO) puede lanzar io.UnsupportedOperation: fileno bajo Werkzeug/Passenger
    return Response(
        data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
