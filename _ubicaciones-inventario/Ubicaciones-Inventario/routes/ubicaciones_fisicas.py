"""
Rutas para gestionar ubicaciones físicas del WMS
"""
from flask import Blueprint, request, jsonify, session
from routes.auth import require_admin
from database import db
from database.models import UbicacionFisica, StockUbicacion
from utils.db_helpers import db_query_with_retry, safe_db_call
from datetime import datetime
import logging
import os
import uuid
from werkzeug.utils import secure_filename

ubicaciones_fisicas_bp = Blueprint('ubicaciones_fisicas', __name__)
logger = logging.getLogger(__name__)

# Configuración para archivos Excel
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
UPLOAD_FOLDER = 'uploads'


@ubicaciones_fisicas_bp.route('/api/ubicaciones-fisicas', methods=['GET'])
@require_admin
def listar_ubicaciones():
    """Lista todas las ubicaciones físicas. Nunca 500 HTML; fallback a lista vacía."""
    request_id = uuid.uuid4().hex[:8]
    logger.info(f"[{request_id}] GET /api/ubicaciones-fisicas")

    def _query():
        return UbicacionFisica.query.order_by(UbicacionFisica.codigo).all()

    result, recovered = safe_db_call(_query, "listar_ubicaciones_fisicas", request_id)

    if result is not None:
        return jsonify({
            "success": True,
            "ubicaciones": [u.to_dict() for u in result],
            "_recovered": recovered
        })

    # Fallback: lista vacía + _fallback (nunca 500 HTML)
    return jsonify({
        "success": True,
        "ubicaciones": [],
        "_fallback": True,
        "error": "Error temporal al conectar con la base de datos",
        "_request_id": request_id
    }), 200


@ubicaciones_fisicas_bp.route('/api/ubicaciones-fisicas', methods=['POST'])
@require_admin
def crear_ubicacion():
    """Crea una nueva ubicación física"""
    try:
        data = request.json or {}
        codigo = data.get('codigo', '').strip().upper()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        tipo = data.get('tipo', '').strip()
        
        # Validaciones
        if not codigo:
            return jsonify({
                "success": False,
                "error": "El código es requerido"
            }), 400
        
        if not nombre:
            return jsonify({
                "success": False,
                "error": "El nombre es requerido"
            }), 400
        
        # Verificar que no exista
        existe = UbicacionFisica.query.filter_by(codigo=codigo).first()
        if existe:
            return jsonify({
                "success": False,
                "error": f"Ya existe una ubicación con el código '{codigo}'"
            }), 400
        
        # Crear ubicación
        ubicacion = UbicacionFisica(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion if descripcion else None,
            tipo=tipo if tipo else None,
            activa=True
        )
        
        db.session.add(ubicacion)
        db.session.commit()
        
        logger.info(f"Ubicación física creada: {codigo} - {nombre}")
        
        return jsonify({
            "success": True,
            "message": f"Ubicación '{codigo}' creada exitosamente",
            "ubicacion": ubicacion.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al crear ubicación física: {e}")
        return jsonify({
            "success": False,
            "error": "Error al crear ubicación física",
            "message": str(e)
        }), 500


@ubicaciones_fisicas_bp.route('/api/ubicaciones-fisicas/<int:ubicacion_id>', methods=['PUT'])
@require_admin
def actualizar_ubicacion(ubicacion_id):
    """Actualiza una ubicación física"""
    try:
        ubicacion = UbicacionFisica.query.get(ubicacion_id)
        if not ubicacion:
            return jsonify({
                "success": False,
                "error": "Ubicación no encontrada"
            }), 404
        
        data = request.json or {}
        codigo = data.get('codigo', '').strip().upper()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        tipo = data.get('tipo', '').strip()
        activa = data.get('activa', True)
        
        # Validaciones
        if not codigo:
            return jsonify({
                "success": False,
                "error": "El código es requerido"
            }), 400
        
        if not nombre:
            return jsonify({
                "success": False,
                "error": "El nombre es requerido"
            }), 400
        
        # Verificar que el código no esté en uso por otra ubicación
        if codigo != ubicacion.codigo:
            existe = UbicacionFisica.query.filter_by(codigo=codigo).first()
            if existe:
                return jsonify({
                    "success": False,
                    "error": f"Ya existe otra ubicación con el código '{codigo}'"
                }), 400
        
        # Actualizar
        ubicacion.codigo = codigo
        ubicacion.nombre = nombre
        ubicacion.descripcion = descripcion if descripcion else None
        ubicacion.tipo = tipo if tipo else None
        ubicacion.activa = activa
        ubicacion.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        logger.info(f"Ubicación física actualizada: {codigo} - {nombre}")
        
        return jsonify({
            "success": True,
            "message": f"Ubicación '{codigo}' actualizada exitosamente",
            "ubicacion": ubicacion.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al actualizar ubicación física: {e}")
        return jsonify({
            "success": False,
            "error": "Error al actualizar ubicación física",
            "message": str(e)
        }), 500


@ubicaciones_fisicas_bp.route('/api/ubicaciones-fisicas/<int:ubicacion_id>', methods=['DELETE'])
@require_admin
def eliminar_ubicacion(ubicacion_id):
    """Elimina una ubicación física (solo si no tiene stock)"""
    try:
        ubicacion = UbicacionFisica.query.get(ubicacion_id)
        if not ubicacion:
            return jsonify({
                "success": False,
                "error": "Ubicación no encontrada"
            }), 404
        
        # Verificar si tiene stock
        stock_count = StockUbicacion.query.filter_by(ubicacion=ubicacion.codigo).count()
        if stock_count > 0:
            return jsonify({
                "success": False,
                "error": f"No se puede eliminar la ubicación '{ubicacion.codigo}' porque tiene {stock_count} productos con stock"
            }), 400
        
        # Eliminar
        db.session.delete(ubicacion)
        db.session.commit()
        
        logger.info(f"Ubicación física eliminada: {ubicacion.codigo}")
        
        return jsonify({
            "success": True,
            "message": f"Ubicación '{ubicacion.codigo}' eliminada exitosamente"
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al eliminar ubicación física: {e}")
        return jsonify({
            "success": False,
            "error": "Error al eliminar ubicación física",
            "message": str(e)
        }), 500


def allowed_file(filename):
    """Verifica si el archivo tiene una extensión permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}


@ubicaciones_fisicas_bp.route('/api/ubicaciones-fisicas/cargar-excel', methods=['POST'])
@require_admin
def cargar_excel():
    """Carga ubicaciones físicas desde un archivo Excel"""
    try:
        if 'archivo' not in request.files:
            return jsonify({
                "success": False,
                "error": "No se proporcionó ningún archivo"
            }), 400
        
        file = request.files['archivo']
        
        if file.filename == '':
            return jsonify({
                "success": False,
                "error": "No se seleccionó ningún archivo"
            }), 400
        
        if not allowed_file(file.filename):
            return jsonify({
                "success": False,
                "error": "Tipo de archivo no permitido. Solo se aceptan archivos .xlsx o .xls"
            }), 400
        
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        max_size = 10 * 1024 * 1024  # 10 MB
        if file_size > max_size:
            return jsonify({
                "success": False,
                "error": f"Archivo demasiado grande ({file_size // 1024}KB). Máximo permitido: 10MB"
            }), 400
        
        # Crear directorio de uploads si no existe
        import os
        UPLOAD_FOLDER = 'uploads'
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)
        
        # Guardar archivo temporalmente
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, f"temp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{filename}")
        file.save(filepath)
        
        try:
            # Intentar importar openpyxl (para .xlsx) o xlrd (para .xls)
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(filepath, data_only=True)
                sheet = workbook.active
                usar_openpyxl = True
            except ImportError:
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(filepath)
                    sheet = workbook.sheet_by_index(0)
                    usar_openpyxl = False
                except ImportError:
                    return jsonify({
                        "success": False,
                        "error": "No se encontró ninguna librería para leer Excel. Por favor instala 'openpyxl' o 'xlrd'"
                    }), 500
            
            # Leer encabezados (primera fila)
            if usar_openpyxl:
                headers = [cell.value for cell in sheet[1] if cell.value]
            else:
                headers = [sheet.cell_value(0, col) for col in range(sheet.ncols) if sheet.cell_value(0, col)]
            
            # Normalizar nombres de columnas (case-insensitive)
            headers_normalizados = {}
            for i, header in enumerate(headers):
                if header:
                    header_str = str(header).strip().upper()
                    headers_normalizados[header_str] = i
            
            # Verificar columnas requeridas
            codigo_col = None
            nombre_col = None
            
            for key in ['CODIGO', 'CÓDIGO', 'CODE']:
                if key in headers_normalizados:
                    codigo_col = headers_normalizados[key]
                    break
            
            for key in ['NOMBRE', 'NAME', 'DESCRIPCION', 'DESCRIPCIÓN']:
                if key in headers_normalizados:
                    nombre_col = headers_normalizados[key]
                    break
            
            if codigo_col is None:
                return jsonify({
                    "success": False,
                    "error": "No se encontró la columna 'Código' en el Excel. Asegúrate de que la primera fila contenga los encabezados."
                }), 400
            
            if nombre_col is None:
                return jsonify({
                    "success": False,
                    "error": "No se encontró la columna 'Nombre' en el Excel. Asegúrate de que la primera fila contenga los encabezados."
                }), 400
            
            # Obtener índices de columnas opcionales
            descripcion_col = headers_normalizados.get('DESCRIPCION') or headers_normalizados.get('DESCRIPCIÓN')
            tipo_col = headers_normalizados.get('TIPO', headers_normalizados.get('TYPE'))
            activa_col = headers_normalizados.get('ACTIVA', headers_normalizados.get('ACTIVE', headers_normalizados.get('ACTIVO')))
            
            # Procesar filas
            creadas = 0
            actualizadas = 0
            errores = []
            total = 0
            
            start_row = 2 if usar_openpyxl else 1  # Empezar desde la fila 2 (después de encabezados)
            max_row = sheet.max_row if usar_openpyxl else sheet.nrows
            
            for row_idx in range(start_row, max_row + 1):
                try:
                    if usar_openpyxl:
                        codigo_val = sheet.cell(row_idx, codigo_col + 1).value
                        nombre_val = sheet.cell(row_idx, nombre_col + 1).value
                        descripcion_val = sheet.cell(row_idx, descripcion_col + 1).value if descripcion_col is not None else None
                        tipo_val = sheet.cell(row_idx, tipo_col + 1).value if tipo_col is not None else None
                        activa_val = sheet.cell(row_idx, activa_col + 1).value if activa_col is not None else None
                    else:
                        codigo_val = sheet.cell_value(row_idx - 1, codigo_col)
                        nombre_val = sheet.cell_value(row_idx - 1, nombre_col)
                        descripcion_val = sheet.cell_value(row_idx - 1, descripcion_col) if descripcion_col is not None else None
                        tipo_val = sheet.cell_value(row_idx - 1, tipo_col) if tipo_col is not None else None
                        activa_val = sheet.cell_value(row_idx - 1, activa_col) if activa_col is not None else None
                    
                    # Validar que código y nombre no estén vacíos
                    if not codigo_val or str(codigo_val).strip() == '':
                        continue  # Saltar filas vacías
                    
                    if not nombre_val or str(nombre_val).strip() == '':
                        errores.append(f"Fila {row_idx}: Código '{codigo_val}' sin nombre")
                        continue
                    
                    # Normalizar valores
                    codigo = str(codigo_val).strip().upper()
                    nombre = str(nombre_val).strip()
                    descripcion = str(descripcion_val).strip() if descripcion_val else None
                    tipo = str(tipo_val).strip() if tipo_val else None
                    activa = True  # Por defecto activa
                    
                    # Procesar campo activa
                    if activa_val is not None:
                        activa_str = str(activa_val).strip().upper()
                        if activa_str in ['FALSE', 'NO', '0', 'INACTIVA', 'INACTIVO']:
                            activa = False
                        elif activa_str in ['TRUE', 'YES', '1', 'SÍ', 'SI', 'ACTIVA', 'ACTIVO']:
                            activa = True
                    
                    # Limpiar valores vacíos
                    if descripcion == '' or descripcion == 'None':
                        descripcion = None
                    if tipo == '' or tipo == 'None':
                        tipo = None
                    
                    total += 1
                    
                    # Buscar si ya existe
                    ubicacion_existente = UbicacionFisica.query.filter_by(codigo=codigo).first()
                    
                    if ubicacion_existente:
                        # Actualizar
                        ubicacion_existente.nombre = nombre
                        ubicacion_existente.descripcion = descripcion
                        ubicacion_existente.tipo = tipo
                        ubicacion_existente.activa = activa
                        ubicacion_existente.updated_at = datetime.utcnow()
                        actualizadas += 1
                        logger.info(f"Ubicación actualizada desde Excel: {codigo} - {nombre}")
                    else:
                        # Crear nueva
                        nueva_ubicacion = UbicacionFisica(
                            codigo=codigo,
                            nombre=nombre,
                            descripcion=descripcion,
                            tipo=tipo,
                            activa=activa
                        )
                        db.session.add(nueva_ubicacion)
                        creadas += 1
                        logger.info(f"Ubicación creada desde Excel: {codigo} - {nombre}")
                
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
                    logger.error(f"Error procesando fila {row_idx} del Excel: {e}")
                    continue
            
            # Commit de todos los cambios
            db.session.commit()
            
            # Eliminar archivo temporal
            try:
                os.remove(filepath)
            except:
                pass
            
            return jsonify({
                "success": True,
                "message": f"Proceso completado: {creadas} creadas, {actualizadas} actualizadas",
                "total": total,
                "creadas": creadas,
                "actualizadas": actualizadas,
                "errores": errores
            })
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error al procesar Excel: {e}")
            # Eliminar archivo temporal en caso de error
            try:
                os.remove(filepath)
            except:
                pass
            return jsonify({
                "success": False,
                "error": f"Error al procesar el archivo Excel: {str(e)}"
            }), 500
    
    except Exception as e:
        logger.error(f"Error en cargar_excel: {e}")
        return jsonify({
            "success": False,
            "error": "Error al cargar el archivo Excel",
            "message": str(e)
        }), 500












